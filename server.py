#!/usr/bin/env python3
"""
Brown Owl Security — Guard Management System
Production | Multi-Admin | Roles | Audit Log | v3.0
Run:  py server.py  →  http://localhost:5000
"""

import http.server, json, sqlite3, os, uuid, base64, re, io, csv, hashlib, secrets, math, time, threading
import smtplib
from email.mime.text import MIMEText
from datetime import datetime, timedelta
from urllib.parse import urlparse, parse_qs

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from PIL import Image as PILImage
    PIL_OK = True
except ImportError:
    PIL_OK = False

# ─── Config ───────────────────────────────────────────────────────────────────
PORT         = int(os.environ.get('PORT', 5000))
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
DATA_DIR     = os.environ.get('DATA_DIR', BASE_DIR)
DB_PATH      = os.path.join(DATA_DIR, 'data', 'security.db')
UPLOADS_PATH = os.path.join(DATA_DIR, 'uploads')
PUBLIC_PATH  = os.path.join(BASE_DIR, 'public')
COMPANY_NAME = os.environ.get('COMPANY_NAME', 'Brown Owl Security (BOS)')
# Default superadmin — used only on first run when no admins exist
DEFAULT_ADMIN_EMAIL    = os.environ.get('ADMIN_EMAIL',    'usamariax0349@gmail.com')
DEFAULT_ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin123')
DEFAULT_ADMIN_NAME     = os.environ.get('ADMIN_NAME',     'Super Admin')
APP_URL      = os.environ.get('APP_URL', f'http://localhost:{int(os.environ.get("PORT", 5000))}')

# ─── SMTP (optional — set env vars to enable email notifications) ─────────────
SMTP_HOST = os.environ.get('SMTP_HOST', '')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASS = os.environ.get('SMTP_PASS', '')
SMTP_FROM = os.environ.get('SMTP_FROM', SMTP_USER)

os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
os.makedirs(UPLOADS_PATH, exist_ok=True)

sessions = {}         # token → {admin_id, role, name, email, created_at, last_seen, ...}
guard_sessions = {}   # token → {guard_id, name, email, created_at, last_seen, ...} — separate privilege domain from admin sessions

# ─── Session expiry ───────────────────────────────────────────────────────────
# Every session used to live in memory forever until an explicit logout or a
# server restart — a leaked token (lost phone, shared computer) stayed valid
# indefinitely. SESSION_IDLE_TIMEOUT expires a session that's gone quiet for a
# day; SESSION_ABSOLUTE_TIMEOUT caps a session at 30 days even if it's in
# constant use, forcing an eventual re-login either way.
SESSION_IDLE_TIMEOUT     = 24 * 3600
SESSION_ABSOLUTE_TIMEOUT = 30 * 86400

def live_session(store, token):
    """Look up `token` in `store` (sessions or guard_sessions), evicting and
    returning None if it's expired, else refreshing last_seen and returning it."""
    s = store.get(token)
    if not s: return None
    now = time.time()
    created  = s.get('created_at', now)
    lastseen = s.get('last_seen', created)
    if now - lastseen > SESSION_IDLE_TIMEOUT or now - created > SESSION_ABSOLUTE_TIMEOUT:
        store.pop(token, None)
        return None
    s['last_seen'] = now
    return s

# ─── Login rate limiting ──────────────────────────────────────────────────────
# In-memory and per-IP (via X-Forwarded-For, since Railway's proxy means
# client_address is never the real caller) rather than per-account, so a
# failed-login flood can't be used to lock a real user out of their own
# account. Resets on restart — the goal is slowing down casual/scripted
# brute force, not maintaining a durable ban list.
LOGIN_MAX_ATTEMPTS  = 10
LOGIN_WINDOW_SECONDS = 15 * 60
_login_failures = {}   # ip → [timestamps of recent failed attempts]

_last_failure_sweep = time.time()
FAILURE_SWEEP_INTERVAL = 3600  # prune stale IP entries at most this often

def _prune_stale_login_failures():
    global _last_failure_sweep
    now = time.time()
    if now - _last_failure_sweep < FAILURE_SWEEP_INTERVAL:
        return
    _last_failure_sweep = now
    for ip in [ip for ip, attempts in _login_failures.items()
               if not attempts or now - attempts[-1] > LOGIN_WINDOW_SECONDS]:
        _login_failures.pop(ip, None)

def login_rate_limited(ip):
    _prune_stale_login_failures()
    now = time.time()
    attempts = [t for t in _login_failures.get(ip, []) if now - t < LOGIN_WINDOW_SECONDS]
    _login_failures[ip] = attempts
    return len(attempts) >= LOGIN_MAX_ATTEMPTS

def record_login_failure(ip):
    _login_failures.setdefault(ip, []).append(time.time())

def record_login_success(ip):
    _login_failures.pop(ip, None)

# ─── Photo uploads ────────────────────────────────────────────────────────────
ALLOWED_PHOTO_EXTS = {'jpg','jpeg','png','gif','webp'}
MAX_PHOTO_B64_CHARS = 8 * 1024 * 1024  # ~6MB decoded — generous for a phone camera photo

def save_uploaded_photo(data):
    """Validate and write a client-submitted photo_b64/photo_ext pair, returning
    the saved filename (or None if no photo was sent). Raises ValueError with a
    user-facing message on anything invalid.

    The extension is never trusted as-is: unlike /api/logo (which already
    whitelists), the incident/submission/clock-out upload paths used to build
    the write path directly from data['photo_ext'] with no check at all —
    a value like '../../../etc/whatever' would land in the filesystem write
    path unvalidated. There's also no size limit until now, so a large
    repeated payload could fill the disk.
    """
    if not data.get('photo_b64') or not data.get('photo_ext'):
        return None
    ext = str(data['photo_ext']).lower().lstrip('.')
    if ext not in ALLOWED_PHOTO_EXTS:
        raise ValueError('Unsupported photo type')
    if len(data['photo_b64']) > MAX_PHOTO_B64_CHARS:
        raise ValueError('Photo is too large')
    photo = f"{uuid.uuid4()}.{ext}"
    with open(os.path.join(UPLOADS_PATH, photo), 'wb') as f:
        f.write(base64.b64decode(data['photo_b64']))
    return photo

# ─── Database backups ─────────────────────────────────────────────────────────
# Rotating on-volume snapshots via sqlite3's own backup API (atomic and safe
# even if a write is mid-transaction — unlike copying the file bytes directly).
# This alone doesn't protect against total volume loss, since the snapshots
# live on the same volume as the live DB — that's what the superadmin-only
# "Download Backup" endpoint further down is for, so a copy can be pulled
# off-Railway too.
BACKUPS_PATH = os.path.join(DATA_DIR, 'data', 'backups')
os.makedirs(BACKUPS_PATH, exist_ok=True)
BACKUP_INTERVAL_SECONDS = 12 * 3600
BACKUP_KEEP = 14

def run_db_backup():
    stamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    dest = os.path.join(BACKUPS_PATH, f'security_{stamp}.db')
    src = sqlite3.connect(DB_PATH)
    dst = sqlite3.connect(dest)
    try:
        src.backup(dst)
    finally:
        dst.close(); src.close()
    kept = sorted(f for f in os.listdir(BACKUPS_PATH) if f.startswith('security_') and f.endswith('.db'))
    for stale in kept[:-BACKUP_KEEP]:
        try: os.remove(os.path.join(BACKUPS_PATH, stale))
        except OSError: pass
    return dest

def _backup_loop():
    while True:
        time.sleep(BACKUP_INTERVAL_SECONDS)
        try: run_db_backup()
        except Exception as e: print(f'  BACKUP: snapshot failed: {e}')

# ─── Password Hashing ─────────────────────────────────────────────────────────
def hash_password(password, salt=None):
    if salt is None:
        salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 100_000)
    return h.hex(), salt

def verify_password(password, stored_hash, salt):
    h, _ = hash_password(password, salt)
    return h == stored_hash

# ─── Geolocation ──────────────────────────────────────────────────────────────
def haversine_m(lat1, lng1, lat2, lng2):
    """Great-circle distance between two points in metres."""
    R = 6371000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lng2 - lng1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * R * math.asin(math.sqrt(a))

# ─── Shift scheduling / live status ──────────────────────────────────────────
CLOCK_IN_GRACE_MINUTES = 30

def shift_status(row, now=None):
    """Computed live status for a scheduled shift — never stored, always derived."""
    now = now or datetime.now()
    if row.get('cancelled'):
        return 'cancelled'
    if row.get('clock_out_at'):
        return 'completed'
    if row.get('clock_in_at'):
        return 'in_progress'
    try:
        start_dt = datetime.strptime(f"{row['shift_date']} {row['start_time']}", '%Y-%m-%d %H:%M')
    except ValueError:
        return 'scheduled'
    if now > start_dt + timedelta(minutes=CLOCK_IN_GRACE_MINUTES):
        return 'missed'
    return 'scheduled'

def with_shift_status(rows):
    now = datetime.now()
    for r in rows:
        r['status'] = shift_status(r, now)
    return rows

# ─── Email ────────────────────────────────────────────────────────────────────
def send_email(to_email, subject, body_text):
    """Send a plain-text email via SMTP. Silently skips if SMTP is not configured."""
    if not SMTP_HOST or not SMTP_USER:
        print(f"  EMAIL: SMTP not configured — skipping notification to {to_email}")
        return False
    try:
        msg = MIMEText(body_text, 'plain')
        msg['Subject'] = subject
        msg['From']    = SMTP_FROM
        msg['To']      = to_email
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.ehlo()
            server.starttls()
            server.login(SMTP_USER, SMTP_PASS)
            server.send_message(msg)
        print(f"  EMAIL: Sent '{subject}' to {to_email}")
        return True
    except Exception as e:
        print(f"  EMAIL: Failed to send to {to_email}: {e}")
        return False

# ─── Database ─────────────────────────────────────────────────────────────────
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.executescript('''
        CREATE TABLE IF NOT EXISTS admins (
            id           TEXT PRIMARY KEY,
            name         TEXT NOT NULL,
            email        TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            salt         TEXT NOT NULL,
            role                 TEXT DEFAULT 'manager',
            active               INTEGER DEFAULT 1,
            must_change_password INTEGER DEFAULT 0,
            last_login           TEXT,
            created_at           TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE UNIQUE INDEX IF NOT EXISTS admins_email ON admins(email);

        CREATE TABLE IF NOT EXISTS guards (
            id             TEXT PRIMARY KEY,
            name           TEXT NOT NULL,
            license_number TEXT,
            base_rate      REAL DEFAULT 0,
            phone          TEXT,
            email          TEXT,
            notes          TEXT,
            active         INTEGER DEFAULT 1,
            created_at     TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS sites (
            id            TEXT PRIMARY KEY,
            name          TEXT NOT NULL,
            client_name   TEXT NOT NULL,
            address       TEXT,
            default_rate  REAL DEFAULT 0,
            contact_name  TEXT,
            contact_phone TEXT,
            active        INTEGER DEFAULT 1,
            created_at    TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS rates (
            guard_id TEXT,
            site_id  TEXT,
            rate     REAL NOT NULL,
            PRIMARY KEY (guard_id, site_id)
        );

        CREATE TABLE IF NOT EXISTS submissions (
            id             TEXT PRIMARY KEY,
            guard_id       TEXT NOT NULL,
            site_id        TEXT NOT NULL,
            shift_date     TEXT NOT NULL,
            start_time     TEXT NOT NULL,
            end_time       TEXT NOT NULL,
            total_hours    REAL NOT NULL,
            notes          TEXT,
            photo_filename TEXT,
            status         TEXT DEFAULT 'pending',
            admin_note     TEXT,
            reviewed_by    TEXT,
            submitted_at   TEXT DEFAULT CURRENT_TIMESTAMP,
            reviewed_at    TEXT
        );

        CREATE TABLE IF NOT EXISTS reminders (
            id         TEXT PRIMARY KEY,
            guard_id   TEXT NOT NULL,
            message    TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            seen_at    TEXT
        );

        CREATE TABLE IF NOT EXISTS audit_log (
            id         TEXT PRIMARY KEY,
            admin_id   TEXT,
            admin_name TEXT,
            action     TEXT NOT NULL,
            details    TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS checkpoints (
            id         TEXT PRIMARY KEY,
            site_id    TEXT NOT NULL,
            name       TEXT NOT NULL,
            lat        REAL NOT NULL,
            lng        REAL NOT NULL,
            radius_m   INTEGER DEFAULT 40,
            sort_order INTEGER DEFAULT 0,
            active     INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS checkpoint_scans (
            id              TEXT PRIMARY KEY,
            checkpoint_id   TEXT NOT NULL,
            checkpoint_name TEXT,
            guard_id        TEXT NOT NULL,
            site_id         TEXT NOT NULL,
            lat             REAL,
            lng             REAL,
            distance_m      REAL,
            scanned_at      TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS incidents (
            id             TEXT PRIMARY KEY,
            guard_id       TEXT NOT NULL,
            site_id        TEXT NOT NULL,
            type           TEXT NOT NULL,
            description    TEXT,
            photo_filename TEXT,
            lat            REAL,
            lng            REAL,
            status         TEXT DEFAULT 'open',
            admin_note     TEXT,
            reviewed_by    TEXT,
            occurred_at    TEXT DEFAULT CURRENT_TIMESTAMP,
            reviewed_at    TEXT
        );

        CREATE TABLE IF NOT EXISTS client_sites (
            admin_id TEXT NOT NULL,
            site_id  TEXT NOT NULL,
            PRIMARY KEY (admin_id, site_id)
        );

        CREATE TABLE IF NOT EXISTS shifts (
            id                  TEXT PRIMARY KEY,
            guard_id            TEXT NOT NULL,
            site_id             TEXT NOT NULL,
            shift_date          TEXT NOT NULL,
            start_time          TEXT NOT NULL,
            end_time            TEXT NOT NULL,
            position            TEXT DEFAULT '',
            notes               TEXT DEFAULT '',
            clock_in_at         TEXT,
            clock_in_lat        REAL,
            clock_in_lng        REAL,
            clock_in_verified   INTEGER DEFAULT 0,
            clock_out_at        TEXT,
            clock_out_lat       REAL,
            clock_out_lng       REAL,
            clock_out_verified  INTEGER DEFAULT 0,
            submission_id       TEXT,
            cancelled           INTEGER DEFAULT 0,
            created_by          TEXT,
            created_at          TEXT DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS guard_site_prefs (
            guard_id TEXT NOT NULL,
            site_id  TEXT NOT NULL,
            pref     TEXT NOT NULL,
            PRIMARY KEY (guard_id, site_id)
        );

        CREATE TABLE IF NOT EXISTS compliance_items (
            id         TEXT PRIMARY KEY,
            name       TEXT NOT NULL,
            sort_order INTEGER DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS guard_compliance (
            id               TEXT PRIMARY KEY,
            guard_id         TEXT NOT NULL,
            item_id          TEXT NOT NULL,
            checked          INTEGER DEFAULT 0,
            reference_no     TEXT DEFAULT '',
            expiry_date      TEXT,
            reminder_days    INTEGER DEFAULT 60,
            critical         INTEGER DEFAULT 0,
            file_filename    TEXT,
            show_to_customer INTEGER DEFAULT 0,
            UNIQUE(guard_id, item_id)
        );

        CREATE TABLE IF NOT EXISTS guard_leave (
            id         TEXT PRIMARY KEY,
            guard_id   TEXT NOT NULL,
            leave_type TEXT NOT NULL DEFAULT 'Fixed Leave',
            start_date TEXT NOT NULL,
            end_date   TEXT NOT NULL,
            notes      TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );

        -- One row per (guard, licence-or-compliance-item, threshold) reminder
        -- actually sent, so the expiry check never nags the same person about
        -- the same deadline twice.
        CREATE TABLE IF NOT EXISTS expiry_reminders_sent (
            id             TEXT PRIMARY KEY,
            guard_id       TEXT NOT NULL,
            item_key       TEXT NOT NULL,   -- 'license' or a compliance_items.id
            threshold_days INTEGER NOT NULL,
            sent_at        TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(guard_id, item_key, threshold_days)
        );

        -- Two-way admin <-> guard conversation thread, one row per message.
        -- Distinct from `reminders` (one-way, system/admin-broadcast notices):
        -- this is an actual back-and-forth per guard, so guards don't have to
        -- text or call the office for something the app can carry.
        CREATE TABLE IF NOT EXISTS messages (
            id          TEXT PRIMARY KEY,
            guard_id    TEXT NOT NULL,
            sender      TEXT NOT NULL,   -- 'admin' or 'guard'
            sender_name TEXT,
            body        TEXT NOT NULL,
            created_at  TEXT DEFAULT CURRENT_TIMESTAMP,
            read_at     TEXT
        );

        -- Keyword-matched canned answers. When a guard's message matches one,
        -- the FAQ auto-responder replies instantly instead of waiting on
        -- admin, and marks the guard's message as already handled.
        CREATE TABLE IF NOT EXISTS faqs (
            id         TEXT PRIMARY KEY,
            question   TEXT NOT NULL,
            keywords   TEXT NOT NULL,   -- comma-separated, matched case-insensitively as substrings
            answer     TEXT NOT NULL,
            active     INTEGER DEFAULT 1,
            sort_order INTEGER DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
    ''')
    conn.commit()

    # Seed the standard compliance catalog once (id-based, so admins can't
    # accidentally duplicate it across restarts)
    COMPLIANCE_CATALOG = ['Covid-19 Vaccination','CPR Accreditation','Crowd Control Card',
        "Driver's Licence",'Firearms Licence','First Aid Certificate','Responsible Service of Alcohol',
        'VISA','White Card','Working With Children Check']
    if conn.execute('SELECT COUNT(*) FROM compliance_items').fetchone()[0] == 0:
        for i, name in enumerate(COMPLIANCE_CATALOG):
            conn.execute('INSERT INTO compliance_items (id,name,sort_order) VALUES (?,?,?)',
                         (str(uuid.uuid4()), name, i))
        conn.commit()

    # Seed a starter set of FAQ auto-replies once, so the feature is useful
    # out of the box — admin can edit, deactivate or add more from the app.
    # Order matters: match_faq() fires the FIRST keyword hit by sort_order, so
    # a narrow/specific question (e.g. a pay *error*) must sort before a
    # broader one that shares vocabulary (a general pay-day question) or the
    # broad one will always win.
    PAY_ERROR_Q  = "There's an error with my pay"
    PAY_ERROR_KW = 'pay error,wrong pay,underpaid,overpaid,pay issue,pay mistake,incorrect pay'
    PAY_ERROR_A  = ("Sorry to hear that. Please reply here with: (1) how much you received, "
                     "(2) how much you expected to receive, (3) your pay rate, (4) how many hours you worked, "
                     "and (5) which site you worked at — the office will investigate and get back to you.")
    OLD_PAYDAY_Q = 'When do I get paid?'
    OLD_PAYDAY_A = ("Pay is processed from your approved shift submissions. "
                     "Contact the office directly for payslip details.")
    FAQ_STARTERS = [
        ('How do I set my availability?', 'availability,day off,time off,leave,holiday',
         "You can set your own availability anytime from the Guard Portal menu → 'My Availability'. "
         "Pick free, partial or unavailable for any day and the office sees it instantly."),
        ('How do I report an incident?', 'incident,report,emergency',
         "Use the Guard Portal menu → 'Report Incident' to log it with photos and your location. "
         "For anything urgent, please call the office directly."),
        ('How do I clock in or out?', 'clock in,clock out,clocking',
         "Open 'My Shifts' in the Guard Portal and use the Clock In / Clock Out button on your shift — "
         "you'll need location access turned on at the site."),
        (PAY_ERROR_Q, PAY_ERROR_KW, PAY_ERROR_A),
        ('When is payment made?', 'pay day,payday,when do i get paid,when is pay,payment day,when is payment',
         "Pay runs every Friday."),
        ('Can I swap a shift with someone?', 'swap,cover my shift,someone take my shift',
         "Shift swaps aren't self-service yet — message the office here with the date and we'll help arrange cover."),
        ('When is the roster published?', 'roster,schedule published,new roster,when is roster',
         "The roster is published every Thursday."),
        ('How does someone apply for a job?', 'apply,job application,vacancy,hiring,want to apply,friend wants to apply',
         "We don't have an online application link yet — message the office here and we'll help with the application process."),
        ("I'm running late for my shift", "running late,late for shift,will be late",
         "Call the office immediately and message us here so we can notify the site."),
        ("I'm sick and can't make my shift", "sick,can't make my shift,calling in sick,unwell",
         "Call the office as soon as possible, and set yourself Unavailable in My Availability so we know for future shifts too."),
        ('How do I update my bank or contact details?', 'bank details,update my details,change my phone,change address,update bank',
         "Contact the office directly to update your bank or contact details."),
        ('I lost my security licence card', 'lost my licence,lost my card,licence missing,lost licence',
         "Contact the office immediately — you may also need to lodge a police report for a replacement."),
        ('Can I get more shifts or hours?', 'more shifts,more hours,extra shifts,pick up shifts',
         "Let the office know you're after more hours, and make sure your Availability is up to date so we can offer you shifts."),
        ('Which site am I working at?', 'which site,what site am i,where am i working,my site today',
         'Check "My Shifts" in the Guard Portal — it shows the site, address and time for each upcoming shift.'),
    ]
    if conn.execute('SELECT COUNT(*) FROM faqs').fetchone()[0] == 0:
        for i, (q, kw, a) in enumerate(FAQ_STARTERS):
            conn.execute('INSERT INTO faqs (id,question,keywords,answer,sort_order) VALUES (?,?,?,?,?)',
                         (str(uuid.uuid4()), q, kw, a, i))
        conn.commit()
    else:
        # Reconcile a DB seeded before this expanded FAQ set existed: repurpose
        # the old generic pay-day row into the pay-error FAQ (only if an admin
        # hasn't already edited it), then append any of the newer starters
        # that aren't present yet, matched by exact question text so nothing
        # is duplicated and no admin edit is ever overwritten.
        old_payday = conn.execute(
            'SELECT id FROM faqs WHERE question=? AND answer=?', (OLD_PAYDAY_Q, OLD_PAYDAY_A)).fetchone()
        if old_payday and not conn.execute(
                'SELECT 1 FROM faqs WHERE question=?', (PAY_ERROR_Q,)).fetchone():
            conn.execute('UPDATE faqs SET question=?,keywords=?,answer=? WHERE id=?',
                         (PAY_ERROR_Q, PAY_ERROR_KW, PAY_ERROR_A, old_payday[0]))
        for q, kw, a in FAQ_STARTERS:
            if not conn.execute('SELECT 1 FROM faqs WHERE question=?', (q,)).fetchone():
                n = conn.execute('SELECT COUNT(*) FROM faqs').fetchone()[0]
                conn.execute('INSERT INTO faqs (id,question,keywords,answer,sort_order) VALUES (?,?,?,?,?)',
                             (str(uuid.uuid4()), q, kw, a, n))
        conn.commit()

    # ── Schema migration: add any missing columns from older databases ──────────
    migrations = [
        # guards table new columns
        ("guards",      "phone",         "ALTER TABLE guards ADD COLUMN phone TEXT DEFAULT ''"),
        ("guards",      "email",         "ALTER TABLE guards ADD COLUMN email TEXT DEFAULT ''"),
        ("guards",      "notes",         "ALTER TABLE guards ADD COLUMN notes TEXT DEFAULT ''"),
        # sites table new columns
        ("sites",       "contact_name",  "ALTER TABLE sites ADD COLUMN contact_name TEXT DEFAULT ''"),
        ("sites",       "contact_phone", "ALTER TABLE sites ADD COLUMN contact_phone TEXT DEFAULT ''"),
        # submissions table new columns
        ("submissions", "admin_note",    "ALTER TABLE submissions ADD COLUMN admin_note TEXT DEFAULT ''"),
        ("submissions", "reviewed_by",   "ALTER TABLE submissions ADD COLUMN reviewed_by TEXT DEFAULT ''"),
        # admins table
        ("admins", "last_login",           "ALTER TABLE admins ADD COLUMN last_login TEXT"),
        ("admins", "must_change_password", "ALTER TABLE admins ADD COLUMN must_change_password INTEGER DEFAULT 0"),
        # sites table — geofence for GPS sign-in verification
        ("sites", "lat",             "ALTER TABLE sites ADD COLUMN lat REAL"),
        ("sites", "lng",             "ALTER TABLE sites ADD COLUMN lng REAL"),
        ("sites", "geofence_radius", "ALTER TABLE sites ADD COLUMN geofence_radius INTEGER DEFAULT 200"),
        # submissions table — captured sign-in location
        ("submissions", "lat",               "ALTER TABLE submissions ADD COLUMN lat REAL"),
        ("submissions", "lng",               "ALTER TABLE submissions ADD COLUMN lng REAL"),
        ("submissions", "distance_m",        "ALTER TABLE submissions ADD COLUMN distance_m REAL"),
        ("submissions", "location_verified", "ALTER TABLE submissions ADD COLUMN location_verified INTEGER DEFAULT 0"),
        # shifts table — draft/publish workflow for the roster grid.
        # DEFAULT 1 so shifts that already exist (and were already visible to
        # guards) stay visible; new shifts are inserted with published=0 explicitly.
        ("shifts", "published", "ALTER TABLE shifts ADD COLUMN published INTEGER DEFAULT 1"),
        # guards table — security licence detail + scheduling flags
        ("guards", "license_state",         "ALTER TABLE guards ADD COLUMN license_state TEXT DEFAULT ''"),
        ("guards", "license_expiry",        "ALTER TABLE guards ADD COLUMN license_expiry TEXT"),
        ("guards", "license_reminder_days", "ALTER TABLE guards ADD COLUMN license_reminder_days INTEGER DEFAULT 60"),
        ("guards", "license_critical",      "ALTER TABLE guards ADD COLUMN license_critical INTEGER DEFAULT 0"),
        ("guards", "license_file",          "ALTER TABLE guards ADD COLUMN license_file TEXT"),
        ("guards", "hide_on_schedule",      "ALTER TABLE guards ADD COLUMN hide_on_schedule INTEGER DEFAULT 0"),
        ("guards", "no_license_required",   "ALTER TABLE guards ADD COLUMN no_license_required INTEGER DEFAULT 0"),
        # An unavailability record with available_from set is a PARTIAL day:
        # the guard is tied up until that time and free afterwards ("free after
        # 6pm"). Empty/NULL keeps the original meaning — unavailable all day.
        ("guard_leave", "available_from",   "ALTER TABLE guard_leave ADD COLUMN available_from TEXT DEFAULT ''"),
        # Set only when the GUARD themselves last touched their availability
        # from the Guard Portal (per-day edit or the "confirm, no changes"
        # button) — never by an admin edit. Lets the admin board show whether
        # a guard's availability is fresh or stale.
        ("guards", "availability_confirmed_at", "ALTER TABLE guards ADD COLUMN availability_confirmed_at TEXT"),
        # Guard Portal login — mirrors the admin auth columns exactly.
        ("guards", "password_hash",        "ALTER TABLE guards ADD COLUMN password_hash TEXT"),
        ("guards", "salt",                 "ALTER TABLE guards ADD COLUMN salt TEXT"),
        ("guards", "must_change_password", "ALTER TABLE guards ADD COLUMN must_change_password INTEGER DEFAULT 0"),
        ("guards", "last_login",           "ALTER TABLE guards ADD COLUMN last_login TEXT"),
    ]
    existing_cols = {}
    for table, col, sql in migrations:
        if table not in existing_cols:
            existing_cols[table] = {r[1] for r in conn.execute(f'PRAGMA table_info({table})').fetchall()}
        if col not in existing_cols[table]:
            conn.execute(sql)
            existing_cols[table].add(col)
            print(f'  Migrated: added {table}.{col}')
    conn.commit()

    # ── Roster seed for the week of 31 Aug – 6 Sep 2026 ──────────────────────
    # seed_data.py (run by the Procfile before this script, on every startup)
    # already loads the full 108-guard/60-site roster, including almost
    # everyone/everywhere below — so this only adds the names genuinely
    # missing from that master list, matched per-name (idempotent on every
    # restart), then the 2 shifts with a confirmed end time from the source
    # roster screenshots. A few names differ only by a middle name/initial
    # in seed_data.py's list (aliased below to the existing record instead
    # of creating a near-duplicate person).
    NAME_ALIASES = {
        'Ahmed Ilyas':        'Ahmed Khalid Ilyas',
        'Usama Khan Niazi':   'Usama arif Khan Niazi',
        'Justin Elzaibak':    'Justin L Elzaibak',
    }
    NEW_GUARDS = ['Abdullah Sajjad', 'Harman (BOS)', 'Shoaib Sherani', 'Faraz Ahmed Kazi',
                  'Sandeep Singh', 'Aman Verma', 'Vikramjeet Singh', 'Rohan Sharma']
    added_guards = 0
    for name in NEW_GUARDS:
        if not conn.execute('SELECT 1 FROM guards WHERE name=?', (name,)).fetchone():
            conn.execute('INSERT INTO guards (id,name) VALUES (?,?)', (str(uuid.uuid4()), name))
            added_guards += 1
    NEW_SITES = [('Skinny Dog Hotel', 'Prime VIC')]  # matches the client_name seed_data.py uses for its siblings
    added_sites = 0
    for name, client in NEW_SITES:
        if not conn.execute('SELECT 1 FROM sites WHERE name=?', (name,)).fetchone():
            conn.execute('INSERT INTO sites (id,name,client_name) VALUES (?,?,?)',
                         (str(uuid.uuid4()), name, client))
            added_sites += 1
    conn.commit()

    def _resolve_guard_id(name):
        row = conn.execute('SELECT id FROM guards WHERE name=?',
                            (NAME_ALIASES.get(name, name),)).fetchone()
        return row[0] if row else None
    def _resolve_site_id(name):
        row = conn.execute('SELECT id FROM sites WHERE name=?', (name,)).fetchone()
        return row[0] if row else None

    SEED_SHIFTS = [
        ('Usama Riaz',   'Anglers Tavern',   '2026-09-06', '06:30', '10:30'),
        ('Rohan Sharma', 'Skinny Dog Hotel', '2026-09-05', '19:00', '23:00'),
    ]
    # "Required" shifts: end time wasn't confirmed in the source roster (common in
    # hospitality security — the shift runs until the venue closes), so end_time
    # is left blank ('') and the UI shows "Required" instead of a bogus time.
    SEED_SHIFTS_OPEN_ENDED = [
        ('Harman (BOS)',        'Eureka Hotel', '2026-09-04', '20:00'),
        ('Mudassar Habib',      'Eureka Hotel', '2026-09-04', '20:00'),
        ('Usama Riaz',          'Eureka Hotel', '2026-09-05', '20:00'),
        ('Mudassar Habib',      'Eureka Hotel', '2026-09-05', '21:00'),
        ('Usama Riaz',          'Eureka Hotel', '2026-09-05', '21:00'),

        ('Joseph Greige',       'Gardiner Hotel', '2026-09-04', '18:00'),
        ('Ahmed Ilyas',         'Gardiner Hotel', '2026-09-04', '19:00'),
        ('Joseph Greige',       'Gardiner Hotel', '2026-09-05', '18:00'),
        ('Ahmed Ilyas',         'Gardiner Hotel', '2026-09-05', '19:00'),
        ('Georges Zaya',        'Gardiner Hotel', '2026-09-05', '20:00'),

        ('Abdullah Sajjad',     'Apollo Bay', '2026-09-05', '18:00'),

        ('Hilal Isik',          'Hotel Esplanade', '2026-09-04', '22:00'),
        ('Shoaib Sherani',      'Hotel Esplanade', '2026-09-04', '22:00'),
        ('Asad ullah Saleem',   'Hotel Esplanade', '2026-09-05', '22:00'),
        ('Faraz Ahmed Kazi',    'Hotel Esplanade', '2026-09-05', '22:00'),
        ('Hilal Isik',          'Hotel Esplanade', '2026-09-05', '22:00'),
        ('Ishtiyaq Ahmed',      'Hotel Esplanade', '2026-09-05', '22:00'),
        ('Jitender Singh',      'Hotel Esplanade', '2026-09-05', '22:00'),
        ('Shoaib Sherani',      'Hotel Esplanade', '2026-09-05', '22:00'),
        ('Usama Khan Niazi',    'Hotel Esplanade', '2026-09-05', '22:00'),

        ('Saied Shohani',       'Melbourne Public', '2026-08-31', '16:30'),
        ('Usama Riaz',          'Melbourne Public', '2026-08-31', '16:30'),
        ('Mohammad Sultan',     'Melbourne Public', '2026-09-04', '18:00'),

        ('Saied Shohani',       'Public House', '2026-09-04', '18:00'),
        ('Justin Elzaibak',     'Public House', '2026-09-05', '17:30'),
        ('Saied Shohani',       'Public House', '2026-09-05', '18:00'),
        ('Nikhil Goyal',        'Public House', '2026-09-05', '20:00'),

        ('Talha Kolcak',        'RSL on Bell', '2026-09-04', '18:00'),
        ('Talha Kolcak',        'RSL on Bell', '2026-09-05', '18:00'),

        ('Harshdeep Singh',     'Swan Hotel', '2026-09-04', '17:00'),
        ('Rahul Kumar',         'Swan Hotel', '2026-09-04', '22:00'),
        ('Zubair Mohammed',     'Swan Hotel', '2026-09-04', '22:00'),
        ('Rahul Kumar',         'Swan Hotel', '2026-09-05', '18:00'),
        ('Harshdeep Singh',     'Swan Hotel', '2026-09-05', '22:00'),
        ('Zubair Mohammed',     'Swan Hotel', '2026-09-05', '22:00'),

        ('Kartikay Sharma',     'The Continental Sorrento', '2026-09-05', '20:30'),
        ('Ali Hussaini',        'The Continental Sorrento', '2026-09-05', '21:00'),
        ('Usama Riaz',          'The Continental Sorrento', '2026-09-05', '21:00'),
        ('Zamin Rezai',         'The Continental Sorrento', '2026-09-05', '21:30'),

        ('Sandeep Singh',       'The Provincial Hotel', '2026-09-04', '20:00'),
        ('Aman Verma',          'The Provincial Hotel', '2026-09-04', '22:00'),
        ('Vikramjeet Singh',    'The Provincial Hotel', '2026-09-04', '22:00'),
        ('Sandeep Singh',       'The Provincial Hotel', '2026-09-05', '20:00'),
        ('Aman Verma',          'The Provincial Hotel', '2026-09-05', '22:00'),
        ('Vikramjeet Singh',    'The Provincial Hotel', '2026-09-05', '22:00'),

        ('Georges Zaya',        'Skinny Dog Hotel', '2026-09-04', '19:00'),
    ]
    seeded_shifts = 0
    for guard_name, site_name, d, st, et in SEED_SHIFTS + [(g, s, d, st, '') for g, s, d, st in SEED_SHIFTS_OPEN_ENDED]:
        gid, sid = _resolve_guard_id(guard_name), _resolve_site_id(site_name)
        if not gid or not sid:
            print(f'  WARNING: could not seed shift for {guard_name} @ {site_name} — guard or site not found')
            continue
        if not conn.execute('''SELECT 1 FROM shifts WHERE guard_id=? AND site_id=?
                                AND shift_date=? AND start_time=?''', (gid, sid, d, st)).fetchone():
            # published=1: this reflects an already-existing roster guards already
            # know about, not a new draft needing Publish & Notify.
            conn.execute('''INSERT INTO shifts (id,guard_id,site_id,shift_date,start_time,end_time,published)
                            VALUES (?,?,?,?,?,?,1)''', (str(uuid.uuid4()), gid, sid, d, st, et))
            seeded_shifts += 1
    conn.commit()
    print(f'  Roster seed: {added_guards} new guards, {added_sites} new sites, {seeded_shifts} shifts added')

    # Always ensure the superadmin from env vars exists with correct password
    h, salt = hash_password(DEFAULT_ADMIN_PASSWORD)
    existing = conn.execute('SELECT id FROM admins WHERE email=?',
                            (DEFAULT_ADMIN_EMAIL.lower(),)).fetchone()
    if existing:
        conn.execute('''UPDATE admins SET password_hash=?, salt=?, role='superadmin', active=1
                        WHERE email=?''', (h, salt, DEFAULT_ADMIN_EMAIL.lower()))
        print(f'  Superadmin password synced: {DEFAULT_ADMIN_EMAIL}')
    else:
        conn.execute('''INSERT INTO admins (id,name,email,password_hash,salt,role)
                        VALUES (?,?,?,?,?,'superadmin')''',
                     (str(uuid.uuid4()), DEFAULT_ADMIN_NAME, DEFAULT_ADMIN_EMAIL.lower(), h, salt))
        print(f'  Superadmin created: {DEFAULT_ADMIN_EMAIL}')
    conn.commit()

    # Case-insensitive unique email per guard, like admins_email — but only
    # among guards who actually have one on file (most don't yet), so this
    # never blocks on the many existing blank-email rows. Wrapped defensively:
    # if any two guards already share an email (in any case), creating the
    # index would throw and we'd rather log that for admin to clean up than
    # crash startup over it. Runs down here, after every guard-seeding step
    # above, so it sees the full roster rather than whatever existed when
    # schema migrations ran.
    try:
        conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS guards_email ON guards(lower(email)) WHERE email != ''")
        conn.commit()
    except sqlite3.IntegrityError as e:
        print(f'  WARNING: guards_email unique index not created — duplicate guard emails exist: {e}')

    # One-time bootstrap: every guard predates password logins, so give each
    # active one without a password a random temp password now rather than
    # locking 100+ guards out the moment this ships. Printed to the startup
    # log (not emailed — most guards have no email on file yet) so admin can
    # hand them out; admin can also reset any individual guard's password
    # later from the Guards tab, which does the same thing on demand.
    needs_password = conn.execute(
        "SELECT id,name,email FROM guards WHERE active=1 AND (password_hash IS NULL OR password_hash='')").fetchall()
    if needs_password:
        print(f'  Guard login rollout: generating temp passwords for {len(needs_password)} guard(s) —')
        for gid, gname, gemail in needs_password:
            temp_pw = secrets.token_urlsafe(6)
            h, salt = hash_password(temp_pw)
            conn.execute('UPDATE guards SET password_hash=?,salt=?,must_change_password=1 WHERE id=?',
                         (h, salt, gid))
            email_note = gemail or '(no email on file — add one before this guard can log in)'
            print(f'    {gname:<30} {email_note:<40} temp password: {temp_pw}')
        conn.commit()

    conn.close()

    # expiring_items()/check_expiry_reminders() read rows by column name (via
    # RL/R), which needs row_factory=sqlite3.Row — init_db's own bare
    # sqlite3.connect() above doesn't set that, so this runs on its own
    # get_db() connection instead, after the migration/seed work is committed.
    db = get_db()
    sent = check_expiry_reminders(db)
    if sent: print(f'  Expiry check: {sent} new licence/compliance reminder(s) sent')
    db.close()

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def R(row):  return dict(row) if row else None
def RL(rows): return [dict(r) for r in rows]

# Guard rows are frequently sent to the admin browser wholesale (SELECT *) —
# strip the password material before that happens, the same way admins.
# password_hash/salt never leave the server for the admins table.
def no_secrets(g):
    if g:
        g.pop('password_hash', None); g.pop('salt', None)
    return g

def no_secrets_list(rows):
    for g in rows: no_secrets(g)
    return rows

def audit(conn, session, action, details=''):
    conn.execute('INSERT INTO audit_log (id,admin_id,admin_name,action,details) VALUES (?,?,?,?,?)',
                 (str(uuid.uuid4()), session.get('admin_id',''), session.get('name',''),
                  action, details))

DAY_NAMES = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']

def delete_submissions(conn, ids, session):
    """Delete submissions and everything that hangs off them.

    A submission is the pay record behind an invoice line, so removing one is
    not just a row delete:
      - the guard's photo would otherwise be orphaned in uploads/
      - shifts.submission_id would point at a row that no longer exists
      - the audit entry has to name what went, since the row itself is gone
    Returns the number actually deleted.
    """
    deleted = 0
    for sid in ids:
        row = R(conn.execute('''SELECT sub.*, g.name as guard_name, s.name as site_name
                                FROM submissions sub
                                JOIN guards g ON g.id=sub.guard_id
                                JOIN sites  s ON s.id=sub.site_id
                                WHERE sub.id=?''', (sid,)).fetchone())
        if not row: continue
        if row.get('photo_filename'):
            try: os.remove(os.path.join(UPLOADS_PATH, row['photo_filename']))
            except OSError: pass  # already gone, or never written — not worth failing the delete
        conn.execute('UPDATE shifts SET submission_id=NULL WHERE submission_id=?', (sid,))
        conn.execute('DELETE FROM submissions WHERE id=?', (sid,))
        audit(conn, session, 'SUBMISSION_DELETE',
              f"{row['guard_name']} @ {row['site_name']} {row['shift_date']} "
              f"{row['total_hours']}h ({row['status']})")
        deleted += 1
    return deleted

_last_expiry_check = 0  # unix time; module-level throttle, resets on restart — fine, see check site
EXPIRY_THRESHOLDS = (30, 14, 3, 0)  # days out; 0 covers "already expired"

def expiring_items(conn, within_days=None):
    """Every guard licence / compliance item with an expiry date, each tagged
    with its days_left (negative = already expired). within_days filters to
    items due within that many days OR already expired; None returns all.
    Shared by the dashboard panel, the Guards list flag, the report, and the
    reminder check below, so all four agree on exactly the same set."""
    today = datetime.now().date()
    rows = []
    for g in RL(conn.execute('''SELECT id,name,license_number,license_expiry,license_reminder_days
                                FROM guards
                                WHERE active=1 AND no_license_required=0
                                  AND license_expiry IS NOT NULL AND license_expiry<>\'\'''').fetchall()):
        try: exp = datetime.strptime(g['license_expiry'], '%Y-%m-%d').date()
        except ValueError: continue
        rows.append({'guard_id':g['id'], 'guard_name':g['name'], 'item_key':'license',
                     'item_name':f"Security Licence {g['license_number'] or ''}".strip(),
                     'expiry_date':g['license_expiry'], 'reminder_days':g['license_reminder_days'] or 60,
                     'days_left':(exp-today).days})
    for gc in RL(conn.execute('''SELECT gc.item_id, gc.expiry_date, gc.reminder_days,
                                        g.id as guard_id, g.name as guard_name, ci.name as item_name
                                 FROM guard_compliance gc
                                 JOIN guards g ON g.id=gc.guard_id
                                 JOIN compliance_items ci ON ci.id=gc.item_id
                                 WHERE g.active=1 AND gc.checked=1
                                   AND gc.expiry_date IS NOT NULL AND gc.expiry_date<>\'\'''').fetchall()):
        try: exp = datetime.strptime(gc['expiry_date'], '%Y-%m-%d').date()
        except ValueError: continue
        rows.append({'guard_id':gc['guard_id'], 'guard_name':gc['guard_name'], 'item_key':gc['item_id'],
                     'item_name':gc['item_name'], 'expiry_date':gc['expiry_date'],
                     'reminder_days':gc['reminder_days'] or 60, 'days_left':(exp-today).days})
    if within_days is not None:
        rows = [r for r in rows if r['days_left'] <= within_days]
    rows.sort(key=lambda r: r['days_left'])
    return rows

def check_expiry_reminders(conn):
    """Fire the 30/14/3-day-out and expired reminders: an in-app reminder for
    the guard (existing Guard Portal popup) plus an admin email, deduped
    per (guard, item, threshold) via expiry_reminders_sent so nobody gets the
    same warning twice. Safe to call often — most calls find nothing new due.

    Each item fires AT MOST ONE reminder per check: the tightest threshold it
    has actually crossed (e.g. something 5 days overdue is just "expired",
    not "expired" + "3 days" + "14 days" + "30 days" all at once because
    today happens to be the first time anyone checked). If a still-tighter
    threshold is crossed later, that fires as its own, separate reminder.
    """
    sent = 0
    for item in expiring_items(conn):
        bucket = next((t for t in sorted(EXPIRY_THRESHOLDS) if item['days_left'] <= t), None)
        if bucket is None:
            continue  # not within any threshold yet
        already = conn.execute(
            'SELECT 1 FROM expiry_reminders_sent WHERE guard_id=? AND item_key=? AND threshold_days=?',
            (item['guard_id'], item['item_key'], bucket)).fetchone()
        if already:
            continue
        when = f"expired {-item['days_left']} day(s) ago" if item['days_left']<0 \
               else f"expires in {item['days_left']} day(s)" if item['days_left']>0 else 'expires today'
        conn.execute('INSERT INTO reminders (id,guard_id,message) VALUES (?,?,?)',
                    (str(uuid.uuid4()), item['guard_id'],
                     f"{item['item_name']} {when} ({item['expiry_date']}). Please renew and update your details."))
        send_email(DEFAULT_ADMIN_EMAIL, f'{COMPANY_NAME}: {item["item_name"]} {when}',
                   f"{item['guard_name']}'s {item['item_name']} {when} (on {item['expiry_date']}).")
        conn.execute('''INSERT INTO expiry_reminders_sent (id,guard_id,item_key,threshold_days)
                        VALUES (?,?,?,?)''',
                    (str(uuid.uuid4()), item['guard_id'], item['item_key'], bucket))
        sent += 1
    if sent: conn.commit()
    return sent

AVAIL_STALE_DAYS = 14  # matches the threshold the admin board already uses for its freshness flag

def guard_availability_status(conn, guard_id, date):
    """What the guard said about one date, for the Add/Edit Shift modal and
    the server-side conflict check on save. Three real states, not two:
    a leave record says 'off' or 'partial' outright; no record could mean
    the guard confirmed they're free, OR it could mean they've simply never
    told the app anything — those must not be conflated, so 'unknown' covers
    both no-confirmation-ever and a confirmation stale enough not to trust.
    Returns {status, available_from, confirmed_at}."""
    leave = R(conn.execute('''SELECT available_from FROM guard_leave
                              WHERE guard_id=? AND start_date<=? AND end_date>=?
                              LIMIT 1''', (guard_id, date, date)).fetchone())
    guard = R(conn.execute('SELECT availability_confirmed_at FROM guards WHERE id=?', (guard_id,)).fetchone())
    confirmed_at = guard['availability_confirmed_at'] if guard else None
    if leave:
        return {'status': 'partial' if leave['available_from'] else 'off',
                'available_from': leave['available_from'] or None, 'confirmed_at': confirmed_at}
    if confirmed_at:
        days_since = (datetime.now() - datetime.fromisoformat(confirmed_at)).days
        if days_since <= AVAIL_STALE_DAYS:
            return {'status': 'free_confirmed', 'available_from': None, 'confirmed_at': confirmed_at}
    return {'status': 'unknown', 'available_from': None, 'confirmed_at': confirmed_at}

def availability_conflict(conn, guard_id, date, start_time):
    """None if no conflict, else a short human string describing it — used
    both for the audit trail on save and to build the modal's warning."""
    a = guard_availability_status(conn, guard_id, date)
    if a['status'] == 'off':
        return 'guard said Unavailable this day'
    if a['status'] == 'partial' and start_time < a['available_from']:
        return f"guard said Free from {a['available_from']}, shift starts {start_time}"
    return None

def match_faq(conn, body):
    """First active FAQ (by sort_order) whose keyword list has a hit in body,
    matched case-insensitively as a plain substring. None if nothing matches."""
    text = (body or '').lower()
    faqs = RL(conn.execute('SELECT * FROM faqs WHERE active=1 ORDER BY sort_order').fetchall())
    for faq in faqs:
        for kw in faq['keywords'].split(','):
            kw = kw.strip().lower()
            if kw and kw in text:
                return faq
    return None

# ─── Invoice helpers ──────────────────────────────────────────────────────────
def invoice_query(qs):
    params, where = [], ["sub.status='approved'"]
    if qs.get('client_name'): where.append('s.client_name=?');    params.append(qs['client_name'][0])
    if qs.get('site_id'):     where.append('sub.site_id=?');       params.append(qs['site_id'][0])
    if qs.get('date_from'):   where.append('sub.shift_date>=?');   params.append(qs['date_from'][0])
    if qs.get('date_to'):     where.append('sub.shift_date<=?');   params.append(qs['date_to'][0])
    db   = get_db()
    rows = RL(db.execute(f'''
        SELECT sub.*, g.name as guard_name, s.name as site_name, s.client_name,
               COALESCE(r.rate, g.base_rate) as rate
        FROM submissions sub
        JOIN guards g ON g.id=sub.guard_id
        JOIN sites  s ON s.id=sub.site_id
        LEFT JOIN rates r ON r.guard_id=sub.guard_id AND r.site_id=sub.site_id
        WHERE {" AND ".join(where)}
        ORDER BY sub.shift_date ASC, g.name ASC
    ''', params).fetchall())
    db.close()
    for r in rows:
        r['amount'] = round(r['total_hours'] * r['rate'], 2)
    return rows, round(sum(r['amount'] for r in rows), 2)

# ─── PDF Generator ────────────────────────────────────────────────────────────
def make_pdf(rows, client_name, date_from, date_to, total):
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    navy = colors.HexColor('#1a2744'); gold = colors.HexColor('#c9a84c')
    lg   = colors.HexColor('#f1f5f9')
    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20*mm, rightMargin=20*mm,
                              topMargin=20*mm, bottomMargin=20*mm)
    ss   = getSampleStyleSheet()
    def P(txt, **kw): return Paragraph(txt, ParagraphStyle('x', parent=ss['Normal'], **kw))
    story = [
        P(COMPANY_NAME, fontSize=22, textColor=navy, fontName='Helvetica-Bold', spaceAfter=2),
        P('Security Services', fontSize=10, textColor=colors.HexColor('#64748b')),
        Spacer(1, 8*mm),
    ]
    inv_no = f"INV-{datetime.now().strftime('%Y%m%d%H%M')}"
    meta   = Table([['INVOICE', inv_no],['Date:', datetime.now().strftime('%d/%m/%Y')],
                    ['Period:', f"{date_from} — {date_to}"],['Bill To:', client_name]],
                   colWidths=[35*mm, 100*mm])
    meta.setStyle(TableStyle([
        ('FONTNAME',(0,0),(0,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(0,0),14),
        ('TEXTCOLOR',(0,0),(0,0),navy),('FONTNAME',(1,0),(1,0),'Helvetica-Bold'),
        ('FONTSIZE',(1,0),(1,0),14),('TEXTCOLOR',(1,0),(1,0),gold),
        ('FONTNAME',(0,1),(0,-1),'Helvetica-Bold'),('FONTSIZE',(0,1),(0,-1),9),
        ('TEXTCOLOR',(0,1),(0,-1),colors.HexColor('#64748b')),
        ('FONTSIZE',(1,1),(1,-1),10),('TEXTCOLOR',(1,1),(1,-1),navy),
        ('BOTTOMPADDING',(0,0),(-1,-1),3),
    ]))
    story += [meta, Spacer(1,8*mm)]
    hdrs = ['Date','Guard','Site','Start','End','Hrs','Rate','Amount']
    td   = [hdrs] + [[r['shift_date'],r['guard_name'],r['site_name'],
                       r['start_time'],r['end_time'],f"{r['total_hours']:.2f}",
                       f"${r['rate']:.2f}",f"${r['amount']:.2f}"] for r in rows]
    cw   = [22*mm,35*mm,35*mm,16*mm,16*mm,14*mm,20*mm,22*mm]
    t    = Table(td, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),navy),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,0),8),
        ('ALIGN',(0,0),(-1,0),'CENTER'),('BOTTOMPADDING',(0,0),(-1,0),6),
        ('TOPPADDING',(0,0),(-1,0),6),('FONTSIZE',(0,1),(-1,-1),8),
        ('TEXTCOLOR',(0,1),(-1,-1),colors.HexColor('#1e293b')),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,lg]),
        ('ALIGN',(5,1),(-1,-1),'RIGHT'),('BOTTOMPADDING',(0,1),(-1,-1),5),
        ('TOPPADDING',(0,1),(-1,-1),5),
        ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#e2e8f0')),
        ('LINEBELOW',(0,0),(-1,0),1,navy),
    ]))
    story += [t, Spacer(1,6*mm)]
    total_tbl = Table([['','','','','','','TOTAL HOURS:',f"{sum(r['total_hours'] for r in rows):.2f}"],
                        ['','','','','','','TOTAL DUE:',f'${total:.2f} AUD']],
                       colWidths=cw)
    total_tbl.setStyle(TableStyle([
        ('FONTNAME',(6,0),(7,1),'Helvetica-Bold'),('FONTSIZE',(6,0),(7,1),10),
        ('TEXTCOLOR',(6,0),(6,1),colors.HexColor('#64748b')),
        ('TEXTCOLOR',(7,0),(7,0),colors.HexColor('#1a2744')),
        ('TEXTCOLOR',(7,1),(7,1),navy),('FONTSIZE',(7,1),(7,1),12),
        ('ALIGN',(6,0),(7,1),'RIGHT'),
        ('LINEABOVE',(6,0),(7,1),1.5,navy),('TOPPADDING',(0,0),(-1,-1),4),
    ]))
    story += [total_tbl, Spacer(1,10*mm),
              P('Thank you for your business.', fontSize=10, textColor=colors.HexColor('#64748b')),
              P('Payment due within 30 days of invoice date.', fontSize=10,
                textColor=colors.HexColor('#64748b'))]
    doc.build(story)
    buf.seek(0)
    return buf.read()

# ─── Excel Generator ──────────────────────────────────────────────────────────
def make_xlsx(rows, client_name, date_from, date_to, total):
    wb   = Workbook()
    ws   = wb.active
    ws.title = 'Invoice'
    nvf  = PatternFill('solid', fgColor='1A2744')
    grf  = PatternFill('solid', fgColor='F1F5F9')
    whtF = Font(color='FFFFFF', bold=True, size=11)
    thin = Side(style='thin', color='CBD5E1')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center')
    rgt  = Alignment(horizontal='right',  vertical='center')
    mid  = Alignment(vertical='center')
    def cell(r, c, val, font=None, fill=None, align=None, border=None):
        ce = ws.cell(r, c, val)
        if font:   ce.font   = font
        if fill:   ce.fill   = fill
        if align:  ce.alignment = align
        if border: ce.border = border
        return ce
    ws.merge_cells('A1:I1')
    cell(1,1, COMPANY_NAME, font=Font(color='1A2744',bold=True,size=16), align=ctr)
    ws.row_dimensions[1].height = 32
    ws.merge_cells('A2:I2')
    cell(2,1, 'Security Services Invoice', font=Font(color='64748B',size=11), align=ctr)
    ws.row_dimensions[2].height = 20
    ws.append([])
    inv_no = f"INV-{datetime.now().strftime('%Y%m%d%H%M')}"
    for la, va, le, ve in [('Invoice #:', inv_no, 'Date:', datetime.now().strftime('%d/%m/%Y')),
                            ('Period:', f"{date_from or '—'} — {date_to or '—'}",
                             'Bill To:', client_name or 'All Clients')]:
        ws.append([la, va, '', '', le, ve])
        r = ws.max_row
        for c, fnt in [(1, Font(color='64748B',bold=True,size=9)),
                       (2, Font(color='1A2744',size=10)),
                       (5, Font(color='64748B',bold=True,size=9)),
                       (6, Font(color='1A2744',size=10))]:
            ws.cell(r, c).font = fnt
    ws.append([])
    hdrs = ['Date','Guard','Site','Client','Start','End','Hours','Rate (AUD)','Amount (AUD)']
    ws.append(hdrs)
    hr = ws.max_row
    for c in range(1, len(hdrs)+1):
        ws.cell(hr,c).fill   = nvf
        ws.cell(hr,c).font   = whtF
        ws.cell(hr,c).alignment = ctr
        ws.cell(hr,c).border = brd
    ws.row_dimensions[hr].height = 22
    total_hrs = 0
    for i, r in enumerate(rows):
        ws.append([r['shift_date'],r['guard_name'],r['site_name'],r['client_name'],
                   r['start_time'],r['end_time'],
                   round(r['total_hours'],2), round(r['rate'],2), r['amount']])
        total_hrs += r['total_hours']
        dr = ws.max_row
        fl = grf if i % 2 == 1 else None
        for c in range(1, len(hdrs)+1):
            ce = ws.cell(dr, c)
            if fl: ce.fill = fl
            ce.border = brd
            ce.alignment = rgt if c >= 7 else mid
        ws.row_dimensions[dr].height = 18
    top_b = Border(top=Side(style='medium', color='1A2744'))
    for label, val in [('TOTAL HOURS:', f"{total_hrs:.2f}"), ('TOTAL DUE:', f'${total:,.2f} AUD')]:
        ws.append(['','','','','','','',label,val])
        r = ws.max_row
        ws.cell(r,8).font = Font(color='64748B',bold=True,size=11)
        ws.cell(r,9).font = Font(color='1A2744',bold=True,size=12)
        ws.cell(r,8).alignment = rgt; ws.cell(r,9).alignment = rgt
        ws.cell(r,8).border = top_b; ws.cell(r,9).border = top_b
        ws.row_dimensions[r].height = 22
    for c, w in enumerate([12,22,22,16,8,8,8,14,15], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    xb = io.BytesIO(); wb.save(xb); xb.seek(0)
    return xb.read()

# ─── Reports hub ──────────────────────────────────────────────────────────────
def rows_to_csv(headers, rows):
    buf = io.StringIO(); w = csv.writer(buf)
    w.writerow(headers)
    for r in rows: w.writerow(r)
    return ('\ufeff'+buf.getvalue()).encode('utf-8')

def rows_to_xlsx(headers, rows, title='Report'):
    wb = Workbook(); ws = wb.active; ws.title = (title[:31] or 'Report')
    navy = PatternFill('solid', fgColor='1A2744')
    white_bold = Font(color='FFFFFF', bold=True, size=11)
    ws.append(headers)
    for c in range(1, len(headers)+1):
        ce = ws.cell(1, c); ce.fill = navy; ce.font = white_bold
    for r in rows: ws.append(r)
    for i, h in enumerate(headers, 1):
        vals = [str(h)] + [str(r[i-1]) for r in rows]
        ws.column_dimensions[get_column_letter(i)].width = min(max(max(len(v) for v in vals)+2, 10), 40)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

def _date_where(qs, col, where, params):
    if qs.get('date_from'): where.append(f'{col}>=?'); params.append(qs['date_from'][0])
    if qs.get('date_to'):   where.append(f'{col}<=?'); params.append(qs['date_to'][0])

def report_schedule_by_guard(qs):
    where, params = ['sh.cancelled=0'], []
    _date_where(qs, 'sh.shift_date', where, params)
    db = get_db()
    rows = with_shift_status(RL(db.execute(f'''
        SELECT sh.*, g.name as guard_name, s.name as site_name, s.client_name
        FROM shifts sh JOIN guards g ON g.id=sh.guard_id JOIN sites s ON s.id=sh.site_id
        WHERE {" AND ".join(where)} ORDER BY g.name, sh.shift_date, sh.start_time
    ''', params).fetchall()))
    db.close()
    headers = ['Guard','Site','Client','Date','Start','End','Position','Status']
    data = [[r['guard_name'],r['site_name'],r['client_name'],r['shift_date'],
             r['start_time'],r['end_time'] or 'Required',r['position'] or '—',r['status']] for r in rows]
    return headers, data

def report_schedule_by_site(qs):
    where, params = ['sh.cancelled=0'], []
    _date_where(qs, 'sh.shift_date', where, params)
    db = get_db()
    rows = with_shift_status(RL(db.execute(f'''
        SELECT sh.*, g.name as guard_name, s.name as site_name, s.client_name
        FROM shifts sh JOIN guards g ON g.id=sh.guard_id JOIN sites s ON s.id=sh.site_id
        WHERE {" AND ".join(where)} ORDER BY s.name, sh.shift_date, sh.start_time
    ''', params).fetchall()))
    db.close()
    headers = ['Site','Client','Guard','Date','Start','End','Position','Status']
    data = [[r['site_name'],r['client_name'],r['guard_name'],r['shift_date'],
             r['start_time'],r['end_time'] or 'Required',r['position'] or '—',r['status']] for r in rows]
    return headers, data

def report_timesheet_approved(qs):
    where, params = ["sub.status='approved'"], []
    _date_where(qs, 'sub.shift_date', where, params)
    db = get_db()
    rows = RL(db.execute(f'''
        SELECT sub.*, g.name as guard_name, g.license_number, s.name as site_name, s.client_name,
               COALESCE(r.rate, g.base_rate) as rate
        FROM submissions sub
        JOIN guards g ON g.id=sub.guard_id
        JOIN sites  s ON s.id=sub.site_id
        LEFT JOIN rates r ON r.guard_id=sub.guard_id AND r.site_id=sub.site_id
        WHERE {" AND ".join(where)} ORDER BY sub.shift_date, g.name
    ''', params).fetchall())
    db.close()
    headers = ['Date','Guard','Licence #','Site','Client','Start','End','Hours','Rate','Amount','Verified']
    data = [[r['shift_date'], r['guard_name'], r['license_number'] or '—', r['site_name'], r['client_name'],
             r['start_time'], r['end_time'], r['total_hours'], r['rate'],
             round(r['total_hours']*r['rate'],2), 'Yes' if r.get('location_verified') else 'No'] for r in rows]
    return headers, data

def report_timesheet_with_notes(qs):
    where, params = ["sub.status='approved'"], []
    _date_where(qs, 'sub.shift_date', where, params)
    db = get_db()
    rows = RL(db.execute(f'''
        SELECT sub.*, g.name as guard_name, s.name as site_name
        FROM submissions sub JOIN guards g ON g.id=sub.guard_id JOIN sites s ON s.id=sub.site_id
        WHERE {" AND ".join(where)} ORDER BY sub.shift_date, g.name
    ''', params).fetchall())
    db.close()
    headers = ['Date','Guard','Site','Start','End','Hours','Notes']
    data = [[r['shift_date'], r['guard_name'], r['site_name'], r['start_time'], r['end_time'],
             r['total_hours'], r['notes'] or ''] for r in rows]
    return headers, data

def report_activity_slip(qs):
    rows, total = invoice_query(qs)
    headers = ['Date','Guard','Site','Client','Start','End','Hours','Rate','Amount']
    data = [[r['shift_date'],r['guard_name'],r['site_name'],r['client_name'],
             r['start_time'],r['end_time'],r['total_hours'],r['rate'],r['amount']] for r in rows]
    data.append(['','','','','','','','TOTAL', round(total,2)])
    return headers, data

def report_live_operations(qs):
    today = datetime.now().strftime('%Y-%m-%d')
    where, params = ['sh.cancelled=0'], []
    where.append('sh.shift_date>=?'); params.append(qs.get('date_from',[today])[0])
    where.append('sh.shift_date<=?'); params.append(qs.get('date_to',[today])[0])
    db = get_db()
    rows = with_shift_status(RL(db.execute(f'''
        SELECT sh.*, g.name as guard_name, s.name as site_name, s.client_name
        FROM shifts sh JOIN guards g ON g.id=sh.guard_id JOIN sites s ON s.id=sh.site_id
        WHERE {" AND ".join(where)} ORDER BY sh.shift_date, sh.start_time
    ''', params).fetchall()))
    db.close()
    def verify_label(r):
        if r['status'] not in ('in_progress','completed'): return '—'
        ok = r.get('clock_in_verified') and (r['status']!='completed' or r.get('clock_out_verified'))
        return 'Verified' if ok else 'Unverified'
    headers = ['Site','Guard','Position','Status','Start Time','End Time','Verification']
    data = [[r['site_name'], r['guard_name'], r['position'] or '—', r['status'],
             r['clock_in_at'] or '—', r['clock_out_at'] or '—', verify_label(r)] for r in rows]
    return headers, data

def report_guard_listing(qs):
    db = get_db()
    rows = RL(db.execute('SELECT * FROM guards ORDER BY active DESC, name').fetchall())
    db.close()
    headers = ['Name','Licence #','Phone','Email','Base Rate','Status']
    data = [[r['name'], r['license_number'] or '—', r['phone'] or '—', r['email'] or '—',
             r['base_rate'], 'Active' if r['active'] else 'Inactive'] for r in rows]
    return headers, data

def report_licence(qs):
    db = get_db()
    rows = RL(db.execute('SELECT * FROM guards WHERE active=1 ORDER BY name').fetchall())
    db.close()
    headers = ['Name','Licence #','Phone','Licence On File']
    data = [[r['name'], r['license_number'] or '—', r['phone'] or '—',
             'Yes' if r['license_number'] else 'No'] for r in rows]
    return headers, data

def report_licenses_expiring(qs):
    db = get_db()
    rows = expiring_items(db)
    db.close()
    headers = ['Guard','Item','Expiry Date','Days Left','Status']
    data = [[r['guard_name'], r['item_name'], r['expiry_date'], r['days_left'],
             'Expired' if r['days_left']<0 else 'Expires Today' if r['days_left']==0 else 'Upcoming'] for r in rows]
    return headers, data

def report_site_listing(qs):
    db = get_db()
    rows = RL(db.execute('SELECT * FROM sites ORDER BY active DESC, client_name, name').fetchall())
    db.close()
    headers = ['Site','Client','Address','Default Rate','Contact Name','Contact Phone','Status']
    data = [[r['name'], r['client_name'], r['address'] or '—', r['default_rate'],
             r['contact_name'] or '—', r['contact_phone'] or '—',
             'Active' if r['active'] else 'Inactive'] for r in rows]
    return headers, data

def report_checkpoints(qs):
    where, params = [], []
    _date_where(qs, 'cs.scanned_at', where, params)
    wc = ('WHERE '+' AND '.join(where)) if where else ''
    db = get_db()
    rows = RL(db.execute(f'''
        SELECT cs.*, g.name as guard_name, s.name as site_name
        FROM checkpoint_scans cs JOIN guards g ON g.id=cs.guard_id JOIN sites s ON s.id=cs.site_id
        {wc} ORDER BY cs.scanned_at DESC
    ''', params).fetchall())
    db.close()
    headers = ['Date/Time','Guard','Site','Checkpoint','Distance (m)']
    data = [[r['scanned_at'], r['guard_name'], r['site_name'], r['checkpoint_name'] or '—',
             round(r['distance_m']) if r['distance_m'] is not None else '—'] for r in rows]
    return headers, data

def report_incidents(qs):
    where, params = [], []
    _date_where(qs, 'i.occurred_at', where, params)
    wc = ('WHERE '+' AND '.join(where)) if where else ''
    db = get_db()
    rows = RL(db.execute(f'''
        SELECT i.*, g.name as guard_name, s.name as site_name
        FROM incidents i JOIN guards g ON g.id=i.guard_id JOIN sites s ON s.id=i.site_id
        {wc} ORDER BY i.occurred_at DESC
    ''', params).fetchall())
    db.close()
    headers = ['Date/Time','Guard','Site','Type','Description','Status']
    data = [[r['occurred_at'], r['guard_name'], r['site_name'], r['type'],
             (r['description'] or ''), r['status']] for r in rows]
    return headers, data

def report_audit(qs):
    where, params = [], []
    _date_where(qs, 'created_at', where, params)
    wc = ('WHERE '+' AND '.join(where)) if where else ''
    db = get_db()
    rows = RL(db.execute(f'SELECT * FROM audit_log {wc} ORDER BY created_at DESC LIMIT 1000', params).fetchall())
    db.close()
    headers = ['Date/Time','Admin','Action','Details']
    data = [[r['created_at'], r['admin_name'] or '—', r['action'], r['details'] or ''] for r in rows]
    return headers, data

REPORTS = {
    'schedule_by_guard': {'category':'Roster Reports','title':'Schedule By Guard',
        'desc':'Generate schedule grouped by guard','fn':report_schedule_by_guard},
    'schedule_by_site': {'category':'Roster Reports','title':'Schedule By Site',
        'desc':'Generate schedule grouped by site','fn':report_schedule_by_site},
    'timesheet_approved': {'category':'Timesheet Reports','title':'Approved Timesheet',
        'desc':'Generate a listing of approved timesheets with pay details','fn':report_timesheet_approved},
    'timesheet_notes': {'category':'Timesheet Reports','title':'Timesheet With Notes',
        'desc':'Generate a listing of approved timesheets including shift notes','fn':report_timesheet_with_notes},
    'activity_slip': {'category':'Activity Reports','title':'Activity Slip',
        'desc':'Generate approved shifts with pay details','fn':report_activity_slip},
    'live_operations': {'category':'Activity Reports','title':'Live Operations Report',
        'desc':'Generate a listing of shifts in the live operations dashboard','fn':report_live_operations},
    'guard_listing': {'category':'Staff Reports','title':'Staff Listing',
        'desc':'Generate a listing of all guards','fn':report_guard_listing},
    'licence': {'category':'Staff Reports','title':'Security Licence Report',
        'desc':'Generate a report for security licences on file','fn':report_licence},
    'licenses_expiring': {'category':'Staff Reports','title':'Licences & Compliance Expiring',
        'desc':'Every guard licence or compliance item with an expiry date, soonest first','fn':report_licenses_expiring},
    'site_listing': {'category':'Site Reports','title':'Site Listing',
        'desc':'Generate a listing of all sites and their contacts','fn':report_site_listing},
    'checkpoints': {'category':'Checkpoint Reports','title':'Checkpoint Report',
        'desc':'Generate a listing of all patrol checkpoint scans','fn':report_checkpoints},
    'incidents': {'category':'Incident Reports','title':'Incident Report',
        'desc':'Generate a listing of reported incidents','fn':report_incidents},
    'invoice_report': {'category':'Invoice Reports','title':'Invoice Report',
        'desc':'Generate invoice-ready billing details, optionally filtered by client','fn':report_activity_slip},
    'audit': {'category':'Audit Reports','title':'Audit Log',
        'desc':'Generate a listing of all admin actions','fn':report_audit},
}

# ─── Handler ──────────────────────────────────────────────────────────────────
class Handler(http.server.BaseHTTPRequestHandler):
    def log_message(self, *a): pass

    def security_headers(self):
        # Cheap defense-in-depth that costs nothing functionally: this app has
        # no legitimate reason to be framed by another site, and browsers
        # should never guess a response's content-type from its bytes. A full
        # script-src CSP isn't attempted here — the frontend leans on inline
        # onclick= handlers throughout, so a strict CSP would break the app
        # wholesale without a much larger refactor (moving every handler to
        # addEventListener). HSTS is safe to send unconditionally since this
        # app should only ever be reached over HTTPS in production.
        self.send_header('X-Content-Type-Options', 'nosniff')
        self.send_header('X-Frame-Options', 'DENY')
        self.send_header('Strict-Transport-Security', 'max-age=31536000; includeSubDomains')

    def send_json(self, data, status=200):
        b = json.dumps(data).encode()
        self.send_response(status)
        self.send_header('Content-Type','application/json')
        self.send_header('Content-Length', len(b))
        self.security_headers()
        self.end_headers(); self.wfile.write(b)

    def err(self, msg, status=400): self.send_json({'error': msg}, status)

    def read_json(self):
        n = int(self.headers.get('Content-Length', 0))
        return json.loads(self.rfile.read(n)) if n else {}

    def client_ip(self):
        # Railway (and any reverse proxy) terminates the real connection, so
        # client_address is the proxy's own address for every request — the
        # actual caller's IP is forwarded in this header instead.
        xff = self.headers.get('X-Forwarded-For')
        return xff.split(',')[0].strip() if xff else self.client_address[0]

    def get_session(self):
        return live_session(sessions, self.headers.get('X-Auth-Token',''))

    def require_admin(self, min_role=None):
        s = self.get_session()
        if not s: self.err('Unauthorized', 401); return None
        # Client accounts are read-only, site-scoped, and never fall through to admin routes
        if s.get('role') == 'client':
            self.err('Insufficient permissions', 403); return None
        role_rank = {'viewer':0,'manager':1,'administrator':2,'superadmin':3}
        if min_role and role_rank.get(s.get('role',''),0) < role_rank.get(min_role,0):
            self.err('Insufficient permissions', 403); return None
        return s

    def require_client(self):
        s = self.get_session()
        if not s: self.err('Unauthorized', 401); return None
        if s.get('role') != 'client':
            self.err('Insufficient permissions', 403); return None
        if s.get('must_change_password'):
            self.err('Please set your password before continuing', 403); return None
        return s

    def get_guard_session(self):
        return live_session(guard_sessions, self.headers.get('X-Auth-Token',''))

    def require_guard(self):
        gs = self.get_guard_session()
        if not gs: self.err('Unauthorized', 401); return None
        return gs

    def client_site_ids(self, admin_id):
        db = get_db()
        ids = [r['site_id'] for r in db.execute(
            'SELECT site_id FROM client_sites WHERE admin_id=?', (admin_id,)).fetchall()]
        db.close(); return ids

    def serve_file(self, path, ct):
        if not os.path.exists(path):
            self.send_response(404); self.end_headers(); return
        with open(path,'rb') as f: data = f.read()
        self.send_response(200)
        self.send_header('Content-Type', ct)
        self.send_header('Content-Length', len(data))
        self.security_headers()
        self.end_headers(); self.wfile.write(data)

    def send_download(self, data, ct, fname):
        self.send_response(200)
        self.send_header('Content-Type', ct)
        self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
        self.send_header('Content-Length', len(data))
        self.security_headers()
        self.end_headers(); self.wfile.write(data)

    # ── GET ────────────────────────────────────────────────────────────────────
    def do_GET(self):
        p = urlparse(self.path); path = p.path; qs = parse_qs(p.query)

        # ── Static ──
        if path in ('/','/index.html'):
            self.serve_file(os.path.join(PUBLIC_PATH,'index.html'),'text/html'); return
        if path == '/manifest.json':
            self.serve_file(os.path.join(PUBLIC_PATH,'manifest.json'),'application/manifest+json'); return
        if path == '/sw.js':
            self.serve_file(os.path.join(PUBLIC_PATH,'sw.js'),'application/javascript'); return
        if path == '/icon.svg':
            svg = '''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 512 512">
<rect width="512" height="512" rx="100" fill="#1a2744"/>
<circle cx="256" cy="190" r="110" fill="#c9a84c" opacity="0.15"/>
<text x="256" y="245" font-family="Arial Black" font-weight="900" font-size="160" fill="#c9a84c" text-anchor="middle">B</text>
<text x="256" y="345" font-family="Arial" font-weight="700" font-size="60" fill="rgba(255,255,255,0.7)" text-anchor="middle" letter-spacing="8">OWL</text>
<text x="256" y="405" font-family="Arial" font-size="36" fill="rgba(255,255,255,0.35)" text-anchor="middle" letter-spacing="4">SECURITY</text>
</svg>'''
            b = svg.encode()
            self.send_response(200); self.send_header('Content-Type','image/svg+xml')
            self.send_header('Content-Length',len(b)); self.end_headers(); self.wfile.write(b); return
        if path.startswith('/uploads/'):
            fname = os.path.basename(path); fpath = os.path.join(UPLOADS_PATH,fname)
            ext = fname.rsplit('.',1)[-1].lower() if '.' in fname else 'jpg'
            ct  = {'jpg':'image/jpeg','jpeg':'image/jpeg','png':'image/png',
                   'gif':'image/gif','webp':'image/webp'}.get(ext,'application/octet-stream')
            self.serve_file(fpath, ct); return
        if path == '/logo':
            for ext in ['png','jpg','jpeg','gif','webp','svg']:
                fp = os.path.join(UPLOADS_PATH,f'company_logo.{ext}')
                if os.path.exists(fp):
                    ct = {'png':'image/png','jpg':'image/jpeg','jpeg':'image/jpeg',
                          'gif':'image/gif','webp':'image/webp','svg':'image/svg+xml'}.get(ext)
                    self.serve_file(fp, ct); return
            self.send_response(404); self.end_headers(); return

        # ── API ──
        if path == '/api/config':
            self.send_json({'company_name': COMPANY_NAME}); return

        if path == '/api/guards':
            db = get_db()
            self.send_json(no_secrets_list(RL(db.execute('SELECT * FROM guards WHERE active=1 ORDER BY name').fetchall())))
            db.close(); return

        if path == '/api/sites':
            db = get_db()
            self.send_json(RL(db.execute('SELECT * FROM sites WHERE active=1 ORDER BY client_name,name').fetchall()))
            db.close(); return

        if path == '/api/checkpoints':
            site_id = qs.get('site_id',[None])[0]
            if not site_id: self.err('site_id required'); return
            db = get_db()
            self.send_json(RL(db.execute(
                'SELECT * FROM checkpoints WHERE site_id=? AND active=1 ORDER BY sort_order, name',
                (site_id,)).fetchall()))
            db.close(); return

        # ── Client portal (read-only, site-scoped) ──
        if path == '/api/client/me':
            s3 = self.require_client()
            if s3 is None: return
            sites = RL(get_db().execute('''
                SELECT s.* FROM sites s JOIN client_sites cs ON cs.site_id=s.id
                WHERE cs.admin_id=? ORDER BY s.name''', (s3['admin_id'],)).fetchall())
            self.send_json({'id':s3['admin_id'],'name':s3['name'],'email':s3['email'],'sites':sites}); return

        if path == '/api/client/activity':
            s3 = self.require_client()
            if s3 is None: return
            site_ids = self.client_site_ids(s3['admin_id'])
            if not site_ids: self.send_json([]); return
            ph = ','.join('?'*len(site_ids))
            db = get_db()
            scans = RL(db.execute(f'''
                SELECT 'checkpoint' as kind, cs.scanned_at as at, cs.checkpoint_name as label,
                       g.name as guard_name, s.name as site_name, cs.distance_m
                FROM checkpoint_scans cs
                JOIN guards g ON g.id=cs.guard_id
                JOIN sites s ON s.id=cs.site_id
                WHERE cs.site_id IN ({ph}) ORDER BY cs.scanned_at DESC LIMIT 100''', site_ids).fetchall())
            incidents = RL(db.execute(f'''
                SELECT 'incident' as kind, i.occurred_at as at, i.type as label,
                       g.name as guard_name, s.name as site_name, i.status
                FROM incidents i
                JOIN guards g ON g.id=i.guard_id
                JOIN sites s ON s.id=i.site_id
                WHERE i.site_id IN ({ph}) ORDER BY i.occurred_at DESC LIMIT 100''', site_ids).fetchall())
            db.close()
            feed = sorted(scans+incidents, key=lambda r:r['at'], reverse=True)[:100]
            self.send_json(feed); return

        if path == '/api/client/incidents':
            s3 = self.require_client()
            if s3 is None: return
            site_ids = self.client_site_ids(s3['admin_id'])
            if not site_ids: self.send_json([]); return
            ph = ','.join('?'*len(site_ids))
            db = get_db()
            rows = RL(db.execute(f'''
                SELECT i.*, g.name as guard_name, s.name as site_name
                FROM incidents i JOIN guards g ON g.id=i.guard_id JOIN sites s ON s.id=i.site_id
                WHERE i.site_id IN ({ph}) ORDER BY i.occurred_at DESC''', site_ids).fetchall())
            db.close(); self.send_json(rows); return

        if path == '/api/client/submissions':
            s3 = self.require_client()
            if s3 is None: return
            site_ids = self.client_site_ids(s3['admin_id'])
            if not site_ids: self.send_json([]); return
            ph = ','.join('?'*len(site_ids))
            db = get_db()
            rows = RL(db.execute(f'''
                SELECT sub.shift_date, sub.start_time, sub.end_time, sub.total_hours, sub.status,
                       sub.location_verified, g.name as guard_name, s.name as site_name
                FROM submissions sub JOIN guards g ON g.id=sub.guard_id JOIN sites s ON s.id=sub.site_id
                WHERE sub.site_id IN ({ph}) AND sub.status='approved'
                ORDER BY sub.shift_date DESC LIMIT 200''', site_ids).fetchall())
            db.close(); self.send_json(rows); return

        # /api/me is available to any authenticated role, including 'client'
        if path == '/api/me':
            s0 = self.get_session()
            if not s0: self.err('Unauthorized', 401); return
            self.send_json({'id':s0['admin_id'],'name':s0['name'],'email':s0['email'],'role':s0['role'],
                            'must_change_password': s0.get('must_change_password', False)}); return

        # Available even mid-forced-password-change, like /api/me for admins
        if path == '/api/guard/me':
            gs0 = self.get_guard_session()
            if not gs0: self.err('Unauthorized', 401); return
            self.send_json({'id':gs0['guard_id'],'name':gs0['name'],'email':gs0['email'],
                            'must_change_password': gs0.get('must_change_password', False)}); return

        # ── Guard-authenticated: every /api/guard/* GET route lives inside this
        # block so a request for anything else (admin routes included) skips
        # it entirely rather than being incorrectly gated by require_guard() ──
        if path.startswith('/api/guard/'):
            gsx = self.require_guard()
            if gsx is None: return
            if gsx.get('must_change_password'):
                self.err('Please set your password before continuing', 403); return

            if path == '/api/guard/shifts':
                db = get_db()
                rows = RL(db.execute('''
                    SELECT sh.*, s.name as site_name, s.client_name, s.address,
                           s.lat as site_lat, s.lng as site_lng, s.geofence_radius
                    FROM shifts sh
                    JOIN sites s ON s.id=sh.site_id
                    WHERE sh.guard_id=? AND sh.shift_date >= date('now','-1 day')
                          AND sh.cancelled=0 AND sh.published=1
                    ORDER BY sh.shift_date ASC, sh.start_time ASC LIMIT 30
                ''', (gsx['guard_id'],)).fetchall())
                db.close(); self.send_json(with_shift_status(rows)); return

            if path == '/api/guard/submissions':
                db = get_db()
                rows = RL(db.execute('''
                    SELECT sub.shift_date, sub.start_time, sub.end_time, sub.total_hours,
                           sub.status, sub.submitted_at, s.name as site_name, s.client_name
                    FROM submissions sub
                    JOIN sites s ON s.id=sub.site_id
                    WHERE sub.guard_id=?
                    ORDER BY sub.submitted_at DESC LIMIT 8
                ''', (gsx['guard_id'],)).fetchall())
                db.close(); self.send_json(rows); return

            if path == '/api/guard/incidents':
                db = get_db()
                rows = RL(db.execute('''
                    SELECT i.*, s.name as site_name
                    FROM incidents i JOIN sites s ON s.id=i.site_id
                    WHERE i.guard_id=? ORDER BY i.occurred_at DESC LIMIT 20
                ''', (gsx['guard_id'],)).fetchall())
                db.close(); self.send_json(rows); return

            # Own availability for a date range — mirrors the old no-auth version,
            # just scoped to the session's guard_id instead of a client-supplied one.
            if path == '/api/guard/availability':
                date_from = qs.get('date_from',[None])[0]
                date_to   = qs.get('date_to',[None])[0]
                if not date_from or not date_to:
                    self.err('date_from and date_to required'); return
                db = get_db()
                guard = R(db.execute('SELECT availability_confirmed_at FROM guards WHERE id=?',
                                     (gsx['guard_id'],)).fetchone())
                leave = RL(db.execute('''SELECT * FROM guard_leave WHERE guard_id=?
                                         AND start_date<=? AND end_date>=?
                                         ORDER BY start_date''',
                                      (gsx['guard_id'], date_to, date_from)).fetchall())
                db.close()
                self.send_json({'confirmed_at': guard['availability_confirmed_at'] if guard else None,
                                'leave': leave}); return

            if path == '/api/guard/messages':
                db = get_db()
                rows = RL(db.execute(
                    'SELECT * FROM messages WHERE guard_id=? ORDER BY created_at ASC', (gsx['guard_id'],)).fetchall())
                db.close(); self.send_json(rows); return

            if path == '/api/guard/compliance':
                db = get_db()
                rows = RL(db.execute('''
                    SELECT ci.id as item_id, ci.name,
                           COALESCE(gc.checked,0) as checked, COALESCE(gc.reference_no,'') as reference_no,
                           gc.expiry_date, COALESCE(gc.reminder_days,60) as reminder_days,
                           COALESCE(gc.critical,0) as critical,
                           COALESCE(gc.show_to_customer,0) as show_to_customer
                    FROM compliance_items ci
                    LEFT JOIN guard_compliance gc ON gc.item_id=ci.id AND gc.guard_id=?
                    ORDER BY ci.sort_order
                ''', (gsx['guard_id'],)).fetchall())
                db.close(); self.send_json(rows); return

            if path == '/api/guard/reminders':
                db = get_db()
                rows = RL(db.execute(
                    "SELECT * FROM reminders WHERE guard_id=? AND seen_at IS NULL ORDER BY created_at DESC",
                    (gsx['guard_id'],)).fetchall())
                db.close(); self.send_json(rows); return

            self.send_response(404); self.end_headers(); return

        # ── Admin-only below ──
        s = self.require_admin()
        if s is None: return

        # Block must_change_password sessions from all other admin GET endpoints
        if s.get('must_change_password'):
            self.err('Please set your password before continuing', 403); return

        if path == '/api/guards/all':
            db = get_db()
            rows = no_secrets_list(RL(db.execute('SELECT * FROM guards ORDER BY active DESC, name').fetchall()))
            db.close(); self.send_json(rows); return

        m = re.match(r'^/api/guards/([^/]+)/site-prefs$', path)
        if m:
            db = get_db()
            rows = RL(db.execute('SELECT site_id, pref FROM guard_site_prefs WHERE guard_id=?',
                                  (m.group(1),)).fetchall())
            db.close(); self.send_json(rows); return

        m = re.match(r'^/api/guards/([^/]+)/compliance$', path)
        if m:
            db = get_db()
            rows = RL(db.execute('''
                SELECT ci.id as item_id, ci.name,
                       COALESCE(gc.checked,0) as checked, COALESCE(gc.reference_no,'') as reference_no,
                       gc.expiry_date, COALESCE(gc.reminder_days,60) as reminder_days,
                       COALESCE(gc.critical,0) as critical, gc.file_filename,
                       COALESCE(gc.show_to_customer,0) as show_to_customer
                FROM compliance_items ci
                LEFT JOIN guard_compliance gc ON gc.item_id=ci.id AND gc.guard_id=?
                ORDER BY ci.sort_order
            ''', (m.group(1),)).fetchall())
            db.close(); self.send_json(rows); return

        m = re.match(r'^/api/guards/([^/]+)/leave$', path)
        if m:
            db = get_db()
            rows = RL(db.execute('SELECT * FROM guard_leave WHERE guard_id=? ORDER BY start_date DESC',
                                  (m.group(1),)).fetchall())
            db.close(); self.send_json(rows); return

        if path == '/api/leave':
            # All guards' leave overlapping a date range — used by the weekly
            # availability view. Without a range, returns everything.
            date_from = qs.get('date_from',[None])[0]
            date_to   = qs.get('date_to',[None])[0]
            where, params = [], []
            if date_to:   where.append('gl.start_date<=?'); params.append(date_to)
            if date_from: where.append('gl.end_date>=?');   params.append(date_from)
            wc = ('WHERE '+' AND '.join(where)) if where else ''
            db = get_db()
            rows = RL(db.execute(f'''
                SELECT gl.*, g.name as guard_name
                FROM guard_leave gl JOIN guards g ON g.id=gl.guard_id
                {wc} ORDER BY gl.start_date
            ''', params).fetchall())
            db.close(); self.send_json(rows); return

        if path == '/api/dashboard':
            db = get_db()
            # No background scheduler in this app (single long-lived process,
            # no cron) — piggyback the expiry check on dashboard loads instead,
            # throttled so a busy admin doesn't rescan on every click. It also
            # runs once at every server startup (see init_db), so an expiring
            # licence is caught within a day either way.
            global _last_expiry_check
            if time.time() - _last_expiry_check > 6*3600:
                check_expiry_reminders(db)
                _last_expiry_check = time.time()
            today = datetime.now().strftime('%Y-%m-%d')
            month_start = datetime.now().strftime('%Y-%m-01')
            pending  = db.execute("SELECT COUNT(*) FROM submissions WHERE status='pending'").fetchone()[0]
            approved_today = db.execute("SELECT COUNT(*) FROM submissions WHERE status='approved' AND DATE(reviewed_at)=?", (today,)).fetchone()[0]
            total_guards = db.execute("SELECT COUNT(*) FROM guards WHERE active=1").fetchone()[0]
            total_sites  = db.execute("SELECT COUNT(*) FROM sites WHERE active=1").fetchone()[0]
            rev_row = db.execute('''
                SELECT COALESCE(SUM(sub.total_hours * COALESCE(r.rate, g.base_rate)),0)
                FROM submissions sub
                JOIN guards g ON g.id=sub.guard_id
                LEFT JOIN rates r ON r.guard_id=sub.guard_id AND r.site_id=sub.site_id
                WHERE sub.status='approved' AND sub.shift_date >= ?
            ''', (month_start,)).fetchone()[0]
            recent = RL(db.execute('''
                SELECT sub.id, sub.shift_date, sub.total_hours, sub.status, sub.submitted_at,
                       g.name as guard_name, s.name as site_name, s.client_name
                FROM submissions sub
                JOIN guards g ON g.id=sub.guard_id
                JOIN sites  s ON s.id=sub.site_id
                ORDER BY sub.submitted_at DESC LIMIT 15
            ''').fetchall())
            expiring = expiring_items(db, within_days=30)
            db.close()
            self.send_json({'pending': pending, 'approved_today': approved_today,
                            'total_guards': total_guards, 'total_sites': total_sites,
                            'revenue_month': round(rev_row, 2), 'recent': recent,
                            'expiring_licenses': expiring}); return

        if path == '/api/licenses/expiring':
            # within_days omitted = everyone with an expiry date on file, for
            # the Reports export; the dashboard panel always passes 30.
            within = qs.get('within_days',[None])[0]
            db = get_db()
            rows = expiring_items(db, within_days=int(within) if within else None)
            db.close(); self.send_json(rows); return

        m = re.match(r'^/api/guards/([^/]+)/availability-status$', path)
        if m:
            date = qs.get('date',[None])[0]
            if not date: self.err('date required'); return
            db = get_db()
            status = guard_availability_status(db, m.group(1), date)
            db.close(); self.send_json(status); return

        if path == '/api/submissions':
            db = get_db(); where=[]; params=[]
            s = sessions.get(self.headers.get('X-Auth-Token',''), {})
            if qs.get('status'):      where.append('sub.status=?');          params.append(qs['status'][0])
            if qs.get('guard_id'):    where.append('sub.guard_id=?');        params.append(qs['guard_id'][0])
            if qs.get('site_id'):     where.append('sub.site_id=?');         params.append(qs['site_id'][0])
            if qs.get('site_name'):   where.append('s.name LIKE ?');         params.append(f"%{qs['site_name'][0]}%")
            if qs.get('client_name'): where.append('s.client_name LIKE ?'); params.append(f"%{qs['client_name'][0]}%")
            if qs.get('date_from'):   where.append('sub.shift_date>=?');     params.append(qs['date_from'][0])
            if qs.get('date_to'):     where.append('sub.shift_date<=?');     params.append(qs['date_to'][0])
            wc = ('WHERE '+' AND '.join(where)) if where else ''
            rows = RL(db.execute(f'''
                SELECT sub.*, g.name as guard_name, g.license_number,
                       s.name as site_name, s.client_name
                FROM submissions sub
                JOIN guards g ON g.id=sub.guard_id
                JOIN sites  s ON s.id=sub.site_id
                {wc}
                ORDER BY sub.shift_date DESC, sub.submitted_at DESC
            ''', params).fetchall())
            db.close(); self.send_json(rows); return

        if path == '/api/rates':
            db = get_db()
            self.send_json(RL(db.execute('''
                SELECT r.*, g.name as guard_name, s.name as site_name, s.client_name
                FROM rates r
                JOIN guards g ON g.id=r.guard_id
                JOIN sites  s ON s.id=r.site_id
                ORDER BY g.name, s.name
            ''').fetchall()))
            db.close(); return

        if path == '/api/invoice':
            rows, total = invoice_query(qs)
            self.send_json({'rows': rows, 'total': total}); return

        if path == '/api/invoice/pdf':
            rows, total = invoice_query(qs)
            cn = qs.get('client_name',[''])[0]; df = qs.get('date_from',[''])[0]; dt = qs.get('date_to',[''])[0]
            try: data = make_pdf(rows, cn or 'All Clients', df or '—', dt or '—', total)
            except Exception as e:
                self.err(f'PDF failed: {e}. Run: py -m pip install reportlab', 500); return
            fn = f"BOS_invoice_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            db = get_db(); audit(db, s, 'INVOICE_PDF', f'client={cn} {df}~{dt}'); db.commit(); db.close()
            self.send_download(data, 'application/pdf', fn); return

        if path == '/api/invoice/xlsx':
            if not OPENPYXL_OK: self.err('Run: py -m pip install openpyxl', 500); return
            rows, total = invoice_query(qs)
            cn = qs.get('client_name',[''])[0]; df = qs.get('date_from',[''])[0]; dt = qs.get('date_to',[''])[0]
            data = make_xlsx(rows, cn or 'All Clients', df, dt, total)
            fn   = f"BOS_invoice_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            db = get_db(); audit(db, s, 'INVOICE_XLSX', f'client={cn} {df}~{dt}'); db.commit(); db.close()
            self.send_download(data,
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', fn); return

        if path == '/api/invoice/csv':
            rows, total = invoice_query(qs)
            buf = io.StringIO(); w = csv.writer(buf)
            w.writerow(['Date','Guard','Site','Client','Start','End','Hours','Rate','Amount'])
            for r in rows:
                w.writerow([r['shift_date'],r['guard_name'],r['site_name'],r['client_name'],
                             r['start_time'],r['end_time'],r['total_hours'],r['rate'],r['amount']])
            w.writerow([]); w.writerow(['','','','','','','','TOTAL',round(total,2)])
            data = ('\ufeff'+buf.getvalue()).encode('utf-8')
            fn   = f"BOS_invoice_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            self.send_download(data,'text/csv; charset=utf-8', fn); return

        if path == '/api/availability/export':
            # The weekly availability board as a spreadsheet: one row per guard,
            # one column per day, matching what's on screen (including the
            # Working/Free/On Leave filter the user has applied).
            if not OPENPYXL_OK: self.err('Run: py -m pip install openpyxl', 500); return
            date_from = qs.get('date_from',[None])[0]
            if not date_from: self.err('date_from required'); return
            only = qs.get('filter',['all'])[0]
            start = datetime.strptime(date_from, '%Y-%m-%d')
            week = [(start + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(7)]
            # days=ISO,ISO limits the export to the days selected on the board.
            picked = [d for d in qs.get('days',[''])[0].split(',') if d in week]
            days = picked or week
            db = get_db()
            guards = RL(db.execute('SELECT id,name FROM guards WHERE active=1 ORDER BY name').fetchall())
            # Query the full week — `days` may be a subset, so days[6] isn't safe.
            shifts = RL(db.execute('''SELECT sh.*, s.name as site_name FROM shifts sh
                                      JOIN sites s ON s.id=sh.site_id
                                      WHERE sh.shift_date>=? AND sh.shift_date<=?
                                        AND COALESCE(sh.cancelled,0)=0''',
                                   (week[0], week[6])).fetchall())
            leave = RL(db.execute('''SELECT * FROM guard_leave
                                     WHERE start_date<=? AND end_date>=?''',
                                  (week[6], week[0])).fetchall())
            db.close()

            def day_shifts(gid, d):
                return [x for x in shifts if x['guard_id']==gid and x['shift_date']==d]
            def day_leave(gid, d):
                for l in leave:
                    if l['guard_id']==gid and l['start_date'] <= d <= l['end_date']: return l
                return None
            def day_full_leave(gid, d):
                # Matches the board: only a full-day record counts as on leave —
                # "free from 18:00" still leaves the guard rosterable.
                l = day_leave(gid, d)
                return l if (l and not l.get('available_from')) else None
            def cell(gid, d):
                sh = day_shifts(gid, d)
                if sh:
                    return '\n'.join(f"{x['start_time']}-{x['end_time'] or 'Required'} {x['site_name']}" for x in sh)
                lv = day_leave(gid, d)
                if lv: return f"Free from {lv['available_from']}" if lv.get('available_from') else 'On Leave'
                return 'Free'

            # Same rule as the board: working/on-leave means ANY day in scope,
            # free means EVERY day in scope.
            def keep(g):
                if only == 'working': return any(day_shifts(g['id'], d) for d in days)
                if only == 'leave':   return any(day_full_leave(g['id'], d) for d in days)
                if only == 'free':
                    return all(not day_shifts(g['id'], d) and not day_full_leave(g['id'], d) for d in days)
                return True

            rows = [[g['name']] + [cell(g['id'], d) for d in days] for g in guards if keep(g)]
            # Weekday comes from the date, not the loop index — `days` can be a
            # subset, so index 0 isn't necessarily Monday.
            headers = ['Guard'] + [f"{DAY_NAMES[datetime.strptime(d,'%Y-%m-%d').weekday()]} {d[8:10]}/{d[5:7]}"
                                   for d in days]
            out = rows_to_xlsx(headers, rows, 'Availability')
            audit_db = get_db(); audit(audit_db, s, 'AVAILABILITY_EXPORT',
                                      f'{days[0]}~{days[-1]} ({len(days)}d, {only})')
            audit_db.commit(); audit_db.close()
            self.send_download(out, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                               f'BOS_Availability_{days[0]}.xlsx'); return

        if path == '/api/reports/catalog':
            self.send_json([{'id':k,'category':v['category'],'title':v['title'],'desc':v['desc']}
                             for k,v in REPORTS.items()]); return

        if path == '/api/reports/run':
            rid = qs.get('report',[''])[0]
            entry = REPORTS.get(rid)
            if not entry: self.err('Unknown report'); return
            try:
                headers, data = entry['fn'](qs)
            except Exception as e:
                self.err(f'Report failed: {e}', 500); return
            fmt   = qs.get('format',['csv'])[0]
            stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            safe  = re.sub(r'[^A-Za-z0-9]+','_', entry['title']).strip('_')
            if fmt == 'xlsx':
                if not OPENPYXL_OK: self.err('Run: py -m pip install openpyxl', 500); return
                out = rows_to_xlsx(headers, data, entry['title'])
                db = get_db(); audit(db, s, 'REPORT_RUN', f"{entry['title']} xlsx"); db.commit(); db.close()
                self.send_download(out, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                    f'BOS_{safe}_{stamp}.xlsx'); return
            out = rows_to_csv(headers, data)
            db = get_db(); audit(db, s, 'REPORT_RUN', f"{entry['title']} csv"); db.commit(); db.close()
            self.send_download(out, 'text/csv; charset=utf-8', f'BOS_{safe}_{stamp}.csv'); return

        if path == '/api/reminders/all':
            db = get_db()
            self.send_json(RL(db.execute('''
                SELECT r.*, g.name as guard_name FROM reminders r
                LEFT JOIN guards g ON g.id=r.guard_id
                ORDER BY r.created_at DESC LIMIT 200
            ''').fetchall()))
            db.close(); return

        # Inbox: one row per guard who has an active conversation, newest first,
        # with the last message and how many of the guard's are still unread.
        if path == '/api/messages/threads':
            db = get_db()
            rows = RL(db.execute('''
                SELECT g.id as guard_id, g.name as guard_name,
                       (SELECT body FROM messages WHERE guard_id=g.id ORDER BY created_at DESC LIMIT 1) as last_body,
                       (SELECT sender FROM messages WHERE guard_id=g.id ORDER BY created_at DESC LIMIT 1) as last_sender,
                       (SELECT created_at FROM messages WHERE guard_id=g.id ORDER BY created_at DESC LIMIT 1) as last_at,
                       (SELECT COUNT(*) FROM messages WHERE guard_id=g.id AND sender='guard' AND read_at IS NULL) as unread
                FROM guards g
                WHERE EXISTS (SELECT 1 FROM messages m WHERE m.guard_id=g.id)
                ORDER BY last_at DESC
            ''').fetchall())
            db.close(); self.send_json(rows); return

        if path == '/api/messages':
            gid = qs.get('guard_id',[None])[0]
            if not gid: self.err('guard_id required'); return
            db = get_db()
            rows = RL(db.execute(
                'SELECT * FROM messages WHERE guard_id=? ORDER BY created_at ASC', (gid,)).fetchall())
            db.close(); self.send_json(rows); return

        if path == '/api/faqs':
            db = get_db()
            rows = RL(db.execute('SELECT * FROM faqs ORDER BY sort_order').fetchall())
            db.close(); self.send_json(rows); return

        if path == '/api/admins':
            s2 = self.require_admin('administrator')
            if not s2: return
            db = get_db()
            rows = RL(db.execute(
                'SELECT id,name,email,role,active,last_login,created_at FROM admins ORDER BY created_at').fetchall())
            # Administrator cannot see superadmin accounts
            if s2.get('role') == 'administrator':
                rows = [r for r in rows if r.get('role') != 'superadmin']
            for r in rows:
                if r['role'] == 'client':
                    r['site_ids'] = self.client_site_ids(r['id'])
            db.close()
            self.send_json(rows); return

        if path == '/api/audit':
            # Guard actions (logins, clock-in/out, incidents, checkpoint scans,
            # availability changes) share this table so there's one durable
            # record of "who did what, when" for compliance/dispute purposes —
            # but there are 100+ guards clocking in and out daily, so by
            # default they're filtered out here to keep the admin-actions view
            # usable; include_guard=1 brings them back in for when they're
            # actually needed.
            include_guard = qs.get('include_guard',['0'])[0] == '1'
            where = '' if include_guard else "WHERE action NOT LIKE 'GUARD_%'"
            db = get_db()
            self.send_json(RL(db.execute(
                f'SELECT * FROM audit_log {where} ORDER BY created_at DESC LIMIT 300').fetchall()))
            db.close(); return

        if path == '/api/backup/download':
            # Superadmin only — streams a fresh snapshot of the live DB so it can
            # be saved off-Railway. The on-volume rotating snapshots (BACKUPS_PATH)
            # protect against a bad migration or accidental delete, but not
            # against losing the volume itself; this is the only thing that does.
            s = self.require_admin('superadmin')
            if s is None: return
            try:
                path_ = run_db_backup()
            except Exception as e:
                self.err(f'Backup failed: {e}', 500); return
            with open(path_, 'rb') as f: data = f.read()
            db = get_db()
            audit(db, s, 'BACKUP_DOWNLOAD'); db.commit(); db.close()
            stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.send_download(data, 'application/octet-stream', f'bos_backup_{stamp}.db'); return

        if path == '/api/incidents':
            db = get_db(); where=[]; params=[]
            if qs.get('status'):  where.append('i.status=?');  params.append(qs['status'][0])
            if qs.get('site_id'): where.append('i.site_id=?'); params.append(qs['site_id'][0])
            wc = ('WHERE '+' AND '.join(where)) if where else ''
            rows = RL(db.execute(f'''
                SELECT i.*, g.name as guard_name, s.name as site_name, s.client_name
                FROM incidents i
                JOIN guards g ON g.id=i.guard_id
                JOIN sites  s ON s.id=i.site_id
                {wc} ORDER BY i.occurred_at DESC LIMIT 300''', params).fetchall())
            db.close(); self.send_json(rows); return

        if path == '/api/activity':
            db = get_db()
            scans = RL(db.execute('''
                SELECT 'checkpoint' as kind, cs.scanned_at as at, cs.checkpoint_name as label,
                       cs.distance_m, g.name as guard_name, s.name as site_name, s.id as site_id
                FROM checkpoint_scans cs
                JOIN guards g ON g.id=cs.guard_id
                JOIN sites  s ON s.id=cs.site_id
                ORDER BY cs.scanned_at DESC LIMIT 60''').fetchall())
            incidents = RL(db.execute('''
                SELECT 'incident' as kind, i.occurred_at as at, i.type as label,
                       i.status, g.name as guard_name, s.name as site_name, s.id as site_id
                FROM incidents i
                JOIN guards g ON g.id=i.guard_id
                JOIN sites  s ON s.id=i.site_id
                ORDER BY i.occurred_at DESC LIMIT 60''').fetchall())
            signins = RL(db.execute('''
                SELECT 'signin' as kind, sub.submitted_at as at, sub.location_verified,
                       sub.distance_m, g.name as guard_name, s.name as site_name, s.id as site_id
                FROM submissions sub
                JOIN guards g ON g.id=sub.guard_id
                JOIN sites  s ON s.id=sub.site_id
                ORDER BY sub.submitted_at DESC LIMIT 60''').fetchall())
            db.close()
            feed = sorted(scans+incidents+signins, key=lambda r:r['at'], reverse=True)[:80]
            self.send_json(feed); return

        if path == '/api/shifts':
            db = get_db(); where=[]; params=[]
            if qs.get('date_from'): where.append('sh.shift_date>=?'); params.append(qs['date_from'][0])
            if qs.get('date_to'):   where.append('sh.shift_date<=?'); params.append(qs['date_to'][0])
            if qs.get('guard_id'):  where.append('sh.guard_id=?');    params.append(qs['guard_id'][0])
            if qs.get('site_id'):   where.append('sh.site_id=?');     params.append(qs['site_id'][0])
            wc = ('WHERE '+' AND '.join(where)) if where else ''
            rows = RL(db.execute(f'''
                SELECT sh.*, g.name as guard_name, s.name as site_name, s.client_name
                FROM shifts sh
                JOIN guards g ON g.id=sh.guard_id
                JOIN sites  s ON s.id=sh.site_id
                {wc} ORDER BY sh.shift_date DESC, sh.start_time DESC LIMIT 500
            ''', params).fetchall())
            db.close()
            rows = with_shift_status(rows)
            status_f = qs.get('status',[None])[0]
            if status_f: rows = [r for r in rows if r['status']==status_f]
            self.send_json(rows); return

        if path == '/api/shifts/live':
            today = datetime.now().strftime('%Y-%m-%d')
            date_from = qs.get('date_from',[today])[0]
            date_to   = qs.get('date_to',[today])[0]
            db = get_db()
            rows = RL(db.execute('''
                SELECT sh.*, g.name as guard_name, s.name as site_name, s.client_name
                FROM shifts sh
                JOIN guards g ON g.id=sh.guard_id
                JOIN sites  s ON s.id=sh.site_id
                WHERE sh.shift_date>=? AND sh.shift_date<=? AND sh.cancelled=0
                ORDER BY sh.shift_date ASC, sh.start_time ASC
            ''', (date_from, date_to)).fetchall())
            db.close()
            rows = with_shift_status(rows)
            counts = {'scheduled':0,'missed':0,'in_progress':0,'completed':0}
            for r in rows: counts[r['status']] = counts.get(r['status'],0) + 1
            self.send_json({'shifts':rows, 'counts':counts, 'total':len(rows)}); return

        self.send_response(404); self.end_headers()

    # ── POST ───────────────────────────────────────────────────────────────────
    def do_POST(self):
        path = urlparse(self.path).path
        data = self.read_json()

        if path == '/api/login':
            ip = self.client_ip()
            if login_rate_limited(ip):
                self.err('Too many failed login attempts. Please try again in a few minutes.', 429); return
            db  = get_db()
            email_in = data.get('email','').strip().lower()
            row = R(db.execute('SELECT * FROM admins WHERE email=? AND active=1',
                               (email_in,)).fetchone())
            db.close()
            print(f"  LOGIN: email={email_in!r} found={'YES' if row else 'NO'}")
            if row:
                pw_ok = verify_password(data.get('password',''), row['password_hash'], row['salt'])
                print(f"  LOGIN: password_check={'PASS' if pw_ok else 'FAIL'} role={row.get('role')}")
                if not pw_ok:
                    record_login_failure(ip)
                    self.err('Invalid email or password', 401); return
            else:
                record_login_failure(ip)
                self.err('Invalid email or password', 401); return
            record_login_success(ip)
            token = str(uuid.uuid4())
            must_change = bool(row.get('must_change_password', 0))
            now = time.time()
            sessions[token] = {'admin_id':row['id'],'name':row['name'],
                                'email':row['email'],'role':row['role'],
                                'must_change_password': must_change,
                                'created_at': now, 'last_seen': now}
            # First-time login: force password change before granting full access
            if must_change:
                self.send_json({'force_password_change': True, 'token': token,
                                'name': row['name']}); return
            db = get_db()
            db.execute('UPDATE admins SET last_login=? WHERE id=?',
                       (datetime.now().isoformat(), row['id']))
            audit(db, sessions[token], 'LOGIN')
            db.commit(); db.close()
            self.send_json({'token':token,'id':row['id'],'name':row['name'],'role':row['role'],'company':COMPANY_NAME}); return

        if path == '/api/logout':
            s = sessions.pop(self.headers.get('X-Auth-Token',''), None)
            if s:
                db = get_db(); audit(db, s, 'LOGOUT'); db.commit(); db.close()
            self.send_json({'ok':True}); return

        if path == '/api/guard/login':
            ip = self.client_ip()
            if login_rate_limited(ip):
                self.err('Too many failed login attempts. Please try again in a few minutes.', 429); return
            db = get_db()
            email_in = data.get('email','').strip().lower()
            row = R(db.execute('SELECT * FROM guards WHERE lower(email)=? AND active=1',
                               (email_in,)).fetchone())
            db.close()
            if not row or not row.get('password_hash') or \
               not verify_password(data.get('password',''), row['password_hash'], row['salt']):
                record_login_failure(ip)
                self.err('Invalid email or password', 401); return
            record_login_success(ip)
            token = str(uuid.uuid4())
            must_change = bool(row.get('must_change_password', 0))
            now = time.time()
            guard_sessions[token] = {'guard_id':row['id'],'name':row['name'],'email':row['email'],
                                      'must_change_password': must_change,
                                      'created_at': now, 'last_seen': now}
            if must_change:
                self.send_json({'force_password_change': True, 'token': token,
                                'name': row['name']}); return
            db = get_db()
            db.execute('UPDATE guards SET last_login=? WHERE id=?',
                       (datetime.now().isoformat(), row['id']))
            audit(db, {'admin_id':row['id'],'name':row['name']}, 'GUARD_LOGIN')
            db.commit(); db.close()
            self.send_json({'token':token,'id':row['id'],'name':row['name'],'company':COMPANY_NAME}); return

        if path == '/api/guard/logout':
            guard_sessions.pop(self.headers.get('X-Auth-Token',''), None)
            self.send_json({'ok':True}); return

        # First-time / forced password setup — mirrors /api/setup-password but
        # on the guards table and guard_sessions, a separate privilege domain
        if path == '/api/guard/setup-password':
            gs0 = self.get_guard_session()
            if not gs0: self.err('Unauthorized', 401); return
            if not gs0.get('must_change_password'):
                self.err('No password change required for this session', 400); return
            new_pw  = data.get('new_password', '')
            conf_pw = data.get('confirm_password', '')
            if len(new_pw) < 6:
                self.err('Password must be at least 6 characters', 400); return
            if new_pw != conf_pw:
                self.err('Passwords do not match', 400); return
            h, salt = hash_password(new_pw)
            db = get_db()
            db.execute('''UPDATE guards SET password_hash=?, salt=?, must_change_password=0, last_login=?
                          WHERE id=?''', (h, salt, datetime.now().isoformat(), gs0['guard_id']))
            db.commit(); db.close()
            gs0['must_change_password'] = False
            self.send_json({'token': self.headers.get('X-Auth-Token',''),
                            'id': gs0['guard_id'], 'name': gs0['name'], 'company': COMPANY_NAME}); return

        # ── Guard-authenticated below — but only for /api/guard/* paths, so a
        # request for anything else (admin routes included) is left alone to
        # fall through to /api/setup-password and Admin-only below, instead
        # of being incorrectly gated by require_guard() ──
        guard_path = path.startswith('/api/guard/')
        if guard_path:
            gsx = self.require_guard()
            if gsx is None: return
            if gsx.get('must_change_password'):
                self.err('Please set your password before continuing', 403); return

        if path == '/api/guard/change-password':
            current = data.get('current_password','')
            new_pw  = data.get('new_password','')
            if not current or not new_pw: self.err('current and new password required'); return
            db = get_db()
            row = R(db.execute('SELECT * FROM guards WHERE id=?',(gsx['guard_id'],)).fetchone())
            if not verify_password(current, row['password_hash'], row['salt']):
                db.close(); self.err('Current password incorrect', 401); return
            h, salt = hash_password(new_pw)
            db.execute('UPDATE guards SET password_hash=?,salt=? WHERE id=?', (h, salt, gsx['guard_id']))
            db.commit(); db.close()
            self.send_json({'ok':True}); return

        if path == '/api/guard/submissions':
            for f in ['site_id','shift_date','start_time','end_time','total_hours']:
                if not data.get(f): self.err(f'{f} required'); return
            db = get_db()
            site = R(db.execute('SELECT lat,lng,geofence_radius,name FROM sites WHERE id=?',
                                (data['site_id'],)).fetchone())
            lat = lng = dist = None
            location_verified = 0
            if site and site.get('lat') is not None and site.get('lng') is not None:
                if data.get('lat') is None or data.get('lng') is None:
                    db.close()
                    self.err('Location access is required to sign in at this site. Please enable location and try again.', 403); return
                lat, lng = float(data['lat']), float(data['lng'])
                radius = site.get('geofence_radius') or 200
                dist = haversine_m(site['lat'], site['lng'], lat, lng)
                if dist > radius:
                    db.close()
                    self.err(f"You must be within {radius}m of {site['name']} to sign in. "
                             f"You are currently {int(dist)}m away — move closer and try again.", 403); return
                location_verified = 1
            try:
                photo = save_uploaded_photo(data)
            except ValueError as e:
                db.close(); self.err(str(e)); return
            sid = str(uuid.uuid4())
            db.execute('''INSERT INTO submissions
                (id,guard_id,site_id,shift_date,start_time,end_time,total_hours,notes,photo_filename,
                 lat,lng,distance_m,location_verified)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (sid,gsx['guard_id'],data['site_id'],data['shift_date'],
                 data['start_time'],data['end_time'],float(data['total_hours']),
                 data.get('notes',''), photo, lat, lng, dist, location_verified))
            audit(db, {'admin_id':gsx['guard_id'],'name':gsx['name']}, 'GUARD_SUBMISSION',
                  f"{data['shift_date']} {data['start_time']}-{data['end_time']} @ {site['name'] if site else data['site_id']}")
            db.commit(); db.close()
            self.send_json({'id':sid,'message':'Shift submitted successfully!'}, 201); return

        # Guard checks in at a patrol checkpoint, GPS-gated
        if path == '/api/guard/checkpoints/scan':
            for f in ['checkpoint_id','lat','lng']:
                if data.get(f) is None: self.err(f'{f} required'); return
            db = get_db()
            cp = R(db.execute('SELECT * FROM checkpoints WHERE id=? AND active=1',
                              (data['checkpoint_id'],)).fetchone())
            if not cp:
                db.close(); self.err('Checkpoint not found', 404); return
            lat, lng = float(data['lat']), float(data['lng'])
            dist = haversine_m(cp['lat'], cp['lng'], lat, lng)
            radius = cp.get('radius_m') or 40
            if dist > radius:
                db.close()
                self.err(f"You must be within {radius}m of '{cp['name']}' to check in. "
                         f"You are currently {int(dist)}m away.", 403); return
            scan_id = str(uuid.uuid4())
            db.execute('''INSERT INTO checkpoint_scans
                (id,checkpoint_id,checkpoint_name,guard_id,site_id,lat,lng,distance_m)
                VALUES (?,?,?,?,?,?,?,?)''',
                (scan_id, cp['id'], cp['name'], gsx['guard_id'], cp['site_id'], lat, lng, dist))
            audit(db, {'admin_id':gsx['guard_id'],'name':gsx['name']}, 'GUARD_CHECKPOINT_SCAN', cp['name'])
            db.commit(); db.close()
            self.send_json({'ok':True,'id':scan_id,'distance_m':round(dist,1)}, 201); return

        if path == '/api/guard/incidents':
            for f in ['site_id','type']:
                if not data.get(f): self.err(f'{f} required'); return
            try:
                photo = save_uploaded_photo(data)
            except ValueError as e:
                self.err(str(e)); return
            iid = str(uuid.uuid4()); db = get_db()
            db.execute('''INSERT INTO incidents (id,guard_id,site_id,type,description,photo_filename,lat,lng)
                          VALUES (?,?,?,?,?,?,?,?)''',
                       (iid, gsx['guard_id'], data['site_id'], data['type'],
                        data.get('description',''), photo, data.get('lat'), data.get('lng')))
            audit(db, {'admin_id':gsx['guard_id'],'name':gsx['name']}, 'GUARD_INCIDENT_REPORT', data['type'])
            db.commit(); db.close()
            self.send_json({'id':iid,'message':'Incident reported.'}, 201); return

        # Mirrors the admin "Set Availability" modal's three states, but only
        # ever touches this guard's own record for this one date (never a
        # multi-day admin-created block on a different date it might overlap).
        if path == '/api/guard/availability':
            if not data.get('date') or not data.get('mode'):
                self.err('date and mode required'); return
            if data['mode'] not in ('free','partial','off'):
                self.err("mode must be 'free', 'partial', or 'off'"); return
            if data['mode']=='partial' and not data.get('available_from'):
                self.err('available_from required for a partial day'); return
            db = get_db()
            date = data['date']
            # Same simplification the admin editor already makes: the record(s)
            # covering this date are replaced wholesale, not split around it.
            db.execute('DELETE FROM guard_leave WHERE guard_id=? AND start_date<=? AND end_date>=?',
                       (gsx['guard_id'], date, date))
            if data['mode'] != 'free':
                db.execute('''INSERT INTO guard_leave (id,guard_id,leave_type,start_date,end_date,notes,available_from)
                              VALUES (?,?,?,?,?,?,?)''',
                           (str(uuid.uuid4()), gsx['guard_id'],
                            'Partial Availability' if data['mode']=='partial' else 'Fixed Leave',
                            date, date, 'Set by guard from the Guard Portal',
                            data.get('available_from','') if data['mode']=='partial' else ''))
            db.execute('UPDATE guards SET availability_confirmed_at=? WHERE id=?',
                       (datetime.now().isoformat(), gsx['guard_id']))
            audit(db, {'admin_id':gsx['guard_id'],'name':gsx['name']}, 'GUARD_AVAILABILITY_SET',
                  f"{date}: {data['mode']}" + (f" from {data.get('available_from')}" if data['mode']=='partial' else ''))
            db.commit(); db.close()
            self.send_json({'ok':True}); return

        # Just refreshes the "last updated" signal the admin board uses to
        # flag stale data, without writing a leave record.
        if path == '/api/guard/availability/confirm':
            db = get_db()
            db.execute('UPDATE guards SET availability_confirmed_at=? WHERE id=?',
                       (datetime.now().isoformat(), gsx['guard_id']))
            db.commit(); db.close()
            self.send_json({'ok':True}); return

        if path == '/api/guard/messages':
            if not (data.get('body') or '').strip():
                self.err('body required'); return
            db = get_db()
            body = data['body'].strip()
            mid = str(uuid.uuid4())
            db.execute('INSERT INTO messages (id,guard_id,sender,sender_name,body) VALUES (?,?,?,?,?)',
                       (mid, gsx['guard_id'], 'guard', gsx['name'], body))
            # FAQ auto-responder: if the message matches a known question,
            # answer instantly and mark it handled instead of leaving it for
            # admin to triage.
            faq = match_faq(db, body)
            if faq:
                now = datetime.now().isoformat()
                db.execute('UPDATE messages SET read_at=? WHERE id=?', (now, mid))
                db.execute('INSERT INTO messages (id,guard_id,sender,sender_name,body) VALUES (?,?,?,?,?)',
                           (str(uuid.uuid4()), gsx['guard_id'], 'admin', 'FAQ Auto-Reply', faq['answer']))
                audit(db, {'admin_id':'','name':'FAQ Auto-Reply'}, 'FAQ_AUTO_REPLY',
                      f"to {gsx['name']}: matched \"{faq['question']}\"")
            db.commit(); db.close()
            self.send_json({'id':mid,'ok':True,'auto_replied':bool(faq)}, 201); return

        if path == '/api/guard/messages/read':
            db = get_db()
            db.execute("UPDATE messages SET read_at=? WHERE guard_id=? AND sender='admin' AND read_at IS NULL",
                       (datetime.now().isoformat(), gsx['guard_id']))
            db.commit(); db.close()
            self.send_json({'ok':True}); return

        if path == '/api/guard/reminders/seen':
            rid = data.get('id')
            if rid:
                db = get_db()
                db.execute("UPDATE reminders SET seen_at=? WHERE id=? AND guard_id=?",
                           (datetime.now().isoformat(), rid, gsx['guard_id']))
                db.commit(); db.close()
            self.send_json({'ok':True}); return

        # Clock in / out of a scheduled shift, GPS-gated. The shift must
        # actually belong to this session's guard — never trusted from the body.
        m_ci = re.match(r'^/api/guard/shifts/([^/]+)/clock-in$', path)
        if m_ci:
            db = get_db()
            sh = R(db.execute('''SELECT sh.*, s.lat as site_lat, s.lng as site_lng,
                                  s.geofence_radius, s.name as site_name
                                  FROM shifts sh JOIN sites s ON s.id=sh.site_id WHERE sh.id=?''',
                              (m_ci.group(1),)).fetchone())
            if not sh: db.close(); self.err('Shift not found', 404); return
            if sh['guard_id'] != gsx['guard_id']: db.close(); self.err('This shift is not assigned to you', 403); return
            if sh['cancelled']: db.close(); self.err('This shift has been cancelled', 400); return
            if sh['clock_in_at']: db.close(); self.err('Already clocked in to this shift', 400); return
            lat = lng = dist = None
            verified = 0
            if sh['site_lat'] is not None and sh['site_lng'] is not None:
                if data.get('lat') is None or data.get('lng') is None:
                    db.close(); self.err('Location access is required to clock in at this site.', 403); return
                lat, lng = float(data['lat']), float(data['lng'])
                radius = sh['geofence_radius'] or 200
                dist = haversine_m(sh['site_lat'], sh['site_lng'], lat, lng)
                if dist > radius:
                    db.close()
                    self.err(f"You must be within {radius}m of {sh['site_name']} to clock in. "
                             f"You are currently {int(dist)}m away.", 403); return
                verified = 1
            db.execute('''UPDATE shifts SET clock_in_at=?, clock_in_lat=?, clock_in_lng=?, clock_in_verified=?
                          WHERE id=?''', (datetime.now().isoformat(), lat, lng, verified, m_ci.group(1)))
            audit(db, {'admin_id':gsx['guard_id'],'name':gsx['name']}, 'GUARD_CLOCK_IN', sh['site_name'])
            db.commit()
            row = with_shift_status([R(db.execute('''SELECT sh.*, g.name as guard_name, s.name as site_name
                FROM shifts sh JOIN guards g ON g.id=sh.guard_id JOIN sites s ON s.id=sh.site_id
                WHERE sh.id=?''', (m_ci.group(1),)).fetchone())])[0]
            db.close(); self.send_json(row); return

        m_co = re.match(r'^/api/guard/shifts/([^/]+)/clock-out$', path)
        if m_co:
            db = get_db()
            sh = R(db.execute('''SELECT sh.*, s.lat as site_lat, s.lng as site_lng,
                                  s.geofence_radius, s.name as site_name
                                  FROM shifts sh JOIN sites s ON s.id=sh.site_id WHERE sh.id=?''',
                              (m_co.group(1),)).fetchone())
            if not sh: db.close(); self.err('Shift not found', 404); return
            if sh['guard_id'] != gsx['guard_id']: db.close(); self.err('This shift is not assigned to you', 403); return
            if not sh['clock_in_at']: db.close(); self.err('You have not clocked in to this shift yet', 400); return
            if sh['clock_out_at']: db.close(); self.err('Already clocked out of this shift', 400); return
            lat = lng = dist = None
            verified = 0
            if sh['site_lat'] is not None and sh['site_lng'] is not None:
                if data.get('lat') is None or data.get('lng') is None:
                    db.close(); self.err('Location access is required to clock out at this site.', 403); return
                lat, lng = float(data['lat']), float(data['lng'])
                radius = sh['geofence_radius'] or 200
                dist = haversine_m(sh['site_lat'], sh['site_lng'], lat, lng)
                if dist > radius:
                    db.close()
                    self.err(f"You must be within {radius}m of {sh['site_name']} to clock out. "
                             f"You are currently {int(dist)}m away.", 403); return
                verified = 1
            now = datetime.now()
            clock_in_dt = datetime.fromisoformat(sh['clock_in_at'])
            total_hours = round((now - clock_in_dt).total_seconds() / 3600, 2)
            try:
                photo = save_uploaded_photo(data)
            except ValueError as e:
                db.close(); self.err(str(e)); return
            sub_id = str(uuid.uuid4())
            location_verified = 1 if (sh['clock_in_verified'] and verified) else 0
            db.execute('''INSERT INTO submissions
                (id,guard_id,site_id,shift_date,start_time,end_time,total_hours,notes,photo_filename,
                 lat,lng,distance_m,location_verified)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (sub_id, sh['guard_id'], sh['site_id'], sh['shift_date'],
                 clock_in_dt.strftime('%H:%M'), now.strftime('%H:%M'), total_hours,
                 data.get('notes',''), photo, lat, lng, dist, location_verified))
            db.execute('''UPDATE shifts SET clock_out_at=?, clock_out_lat=?, clock_out_lng=?,
                          clock_out_verified=?, submission_id=? WHERE id=?''',
                       (now.isoformat(), lat, lng, verified, sub_id, m_co.group(1)))
            audit(db, {'admin_id':gsx['guard_id'],'name':gsx['name']}, 'GUARD_CLOCK_OUT',
                  f"{sh['site_name']} · {total_hours}h")
            db.commit()
            row = with_shift_status([R(db.execute('''SELECT sh.*, g.name as guard_name, s.name as site_name
                FROM shifts sh JOIN guards g ON g.id=sh.guard_id JOIN sites s ON s.id=sh.site_id
                WHERE sh.id=?''', (m_co.group(1),)).fetchone())])[0]
            db.close(); self.send_json({**row, 'submission_id': sub_id, 'total_hours': total_hours}); return

        if guard_path:
            self.send_response(404); self.end_headers(); return

        # ── First-login password setup — available to any authenticated role ────
        if path == '/api/setup-password':
            s = self.get_session()
            if not s: self.err('Unauthorized', 401); return
            if not s.get('must_change_password'):
                self.err('No password change required for this session', 400); return
            new_pw  = data.get('new_password', '')
            conf_pw = data.get('confirm_password', '')
            if len(new_pw) < 6:
                self.err('Password must be at least 6 characters', 400); return
            if new_pw != conf_pw:
                self.err('Passwords do not match', 400); return
            h, salt = hash_password(new_pw)
            db = get_db()
            db.execute('UPDATE admins SET password_hash=?, salt=?, must_change_password=0 WHERE id=?',
                       (h, salt, s['admin_id']))
            audit(db, s, 'PASSWORD_SETUP_COMPLETE'); db.commit(); db.close()
            # Upgrade session to full access
            s['must_change_password'] = False
            db = get_db()
            db.execute('UPDATE admins SET last_login=? WHERE id=?',
                       (datetime.now().isoformat(), s['admin_id']))
            audit(db, s, 'LOGIN'); db.commit(); db.close()
            self.send_json({'token': self.headers.get('X-Auth-Token',''),
                            'id': s['admin_id'], 'name': s['name'],
                            'role': s['role'], 'company': COMPANY_NAME}); return

        # ── Admin-only below ──
        s = self.require_admin()
        if s is None: return

        # Block must_change_password sessions from all other admin POST endpoints
        if s.get('must_change_password'):
            self.err('Please set your password before continuing', 403); return

        m_rot = re.match(r'^/api/submissions/([^/]+)/rotate$', path)
        if m_rot:
            s2 = self.require_admin('manager')
            if not s2: return
            if not PIL_OK: self.err('Image rotation requires Pillow. Run: pip install Pillow', 500); return
            degrees = int(data.get('degrees', 90))
            if degrees not in (90, 180, 270): self.err('degrees must be 90, 180 or 270'); return
            db = get_db()
            row = R(db.execute('SELECT photo_filename FROM submissions WHERE id=?',
                               (m_rot.group(1),)).fetchone())
            if not row or not row['photo_filename']:
                db.close(); self.err('No photo for this submission'); return
            fpath = os.path.join(UPLOADS_PATH, row['photo_filename'])
            if not os.path.exists(fpath):
                db.close(); self.err('Photo file not found on server'); return
            try:
                img = PILImage.open(fpath)
                rotated = img.rotate(-degrees, expand=True)
                rotated.save(fpath)
                audit(db, s, 'PHOTO_ROTATED', f"sub={m_rot.group(1)} {degrees}deg")
                db.commit()
            except Exception as e:
                db.close(); self.err(f'Rotation failed: {e}', 500); return
            db.close()
            self.send_json({'ok': True, 'filename': row['photo_filename']}); return

        if path == '/api/guards':
            s2 = self.require_admin('manager')
            if not s2: return
            if not data.get('name'): self.err('Name required'); return
            gid = str(uuid.uuid4()); db = get_db()
            try:
                db.execute('''INSERT INTO guards (id,name,license_number,base_rate,phone,email,notes)
                              VALUES (?,?,?,?,?,?,?)''',
                           (gid,data['name'],data.get('license_number',''),
                            float(data.get('base_rate',0)),data.get('phone',''),
                            data.get('email',''),data.get('notes','')))
            except sqlite3.IntegrityError:
                db.close(); self.err('Another guard already has that email address', 409); return
            db.commit()
            audit(db, s, 'GUARD_CREATE', data['name']); db.commit()
            g = no_secrets(R(db.execute('SELECT * FROM guards WHERE id=?',(gid,)).fetchone())); db.close()
            self.send_json(g, 201); return

        # Set/reset a guard's Guard Portal password — same idea as inviting an
        # admin: generate a temp password, force them to replace it on first
        # login, and hand the plaintext back once here (never stored) so it
        # can be emailed and/or shown to the admin to pass along by hand —
        # most guards don't have an email on file yet.
        m = re.match(r'^/api/guards/([^/]+)/reset-password$', path)
        if m:
            s2 = self.require_admin('manager')
            if not s2: return
            db = get_db()
            guard = R(db.execute('SELECT id,name,email FROM guards WHERE id=?', (m.group(1),)).fetchone())
            if not guard: db.close(); self.err('Guard not found', 404); return
            temp_pw = secrets.token_urlsafe(6)
            h, salt = hash_password(temp_pw)
            db.execute('UPDATE guards SET password_hash=?,salt=?,must_change_password=1 WHERE id=?',
                       (h, salt, guard['id']))
            audit(db, s2, 'GUARD_PASSWORD_RESET', guard['name']); db.commit(); db.close()
            emailed = False
            if guard.get('email'):
                emailed = send_email(
                    guard['email'],
                    f"Your {COMPANY_NAME} Guard Portal Password",
                    f"Hi {guard['name']},\n\n"
                    f"A temporary password has been set for your Guard Portal account.\n\n"
                    f"Login URL:          {APP_URL}\n"
                    f"Email:              {guard['email']}\n"
                    f"Temporary password: {temp_pw}\n\n"
                    f"IMPORTANT: You will be prompted to set your own password the first time you log in. "
                    f"Your temporary password will no longer work after that.\n\n"
                    f"— {COMPANY_NAME}"
                )
            self.send_json({'ok':True,'temp_password':temp_pw,'emailed':emailed}); return

        m = re.match(r'^/api/guards/([^/]+)/site-pref$', path)
        if m:
            s2 = self.require_admin('manager')
            if not s2: return
            site_id = data.get('site_id'); pref = data.get('pref') or ''
            if not site_id: self.err('site_id required'); return
            db = get_db()
            if pref in ('blacklist','preferred'):
                db.execute('''INSERT INTO guard_site_prefs (guard_id,site_id,pref) VALUES (?,?,?)
                              ON CONFLICT(guard_id,site_id) DO UPDATE SET pref=excluded.pref''',
                           (m.group(1), site_id, pref))
            else:
                db.execute('DELETE FROM guard_site_prefs WHERE guard_id=? AND site_id=?',
                           (m.group(1), site_id))
            audit(db, s, 'GUARD_SITE_PREF', f'{m.group(1)} site={site_id} pref={pref or "none"}')
            db.commit(); db.close()
            self.send_json({'ok':True}); return

        m = re.match(r'^/api/guards/([^/]+)/blacklist-all$', path)
        if m:
            s2 = self.require_admin('manager')
            if not s2: return
            db = get_db()
            site_ids = [r[0] for r in db.execute('SELECT id FROM sites WHERE active=1').fetchall()]
            for sid in site_ids:
                db.execute('''INSERT INTO guard_site_prefs (guard_id,site_id,pref) VALUES (?,?,'blacklist')
                              ON CONFLICT(guard_id,site_id) DO UPDATE SET pref='blacklist' ''',
                           (m.group(1), sid))
            audit(db, s, 'GUARD_BLACKLIST_ALL', m.group(1)); db.commit(); db.close()
            self.send_json({'ok':True, 'count':len(site_ids)}); return

        m = re.match(r'^/api/guards/([^/]+)/compliance/([^/]+)/file$', path)
        if m:
            s2 = self.require_admin('manager')
            if not s2: return
            if not data.get('file_b64'): self.err('file_b64 required'); return
            ext = re.sub(r'[^a-zA-Z0-9]', '', data.get('ext','dat'))[:6] or 'dat'
            fname = f"compliance_{uuid.uuid4().hex}.{ext}"
            with open(os.path.join(UPLOADS_PATH, fname), 'wb') as f:
                f.write(base64.b64decode(data['file_b64']))
            db = get_db()
            db.execute('''INSERT INTO guard_compliance (id,guard_id,item_id,file_filename) VALUES (?,?,?,?)
                          ON CONFLICT(guard_id,item_id) DO UPDATE SET file_filename=excluded.file_filename''',
                       (str(uuid.uuid4()), m.group(1), m.group(2), fname))
            audit(db, s, 'COMPLIANCE_FILE', f'{m.group(1)} item={m.group(2)}'); db.commit(); db.close()
            self.send_json({'ok':True, 'filename':fname}); return

        m = re.match(r'^/api/guards/([^/]+)/license-file$', path)
        if m:
            s2 = self.require_admin('manager')
            if not s2: return
            if not data.get('file_b64'): self.err('file_b64 required'); return
            ext = re.sub(r'[^a-zA-Z0-9]', '', data.get('ext','dat'))[:6] or 'dat'
            fname = f"license_{uuid.uuid4().hex}.{ext}"
            with open(os.path.join(UPLOADS_PATH, fname), 'wb') as f:
                f.write(base64.b64decode(data['file_b64']))
            db = get_db()
            db.execute('UPDATE guards SET license_file=? WHERE id=?', (fname, m.group(1)))
            audit(db, s, 'LICENSE_FILE', m.group(1)); db.commit(); db.close()
            self.send_json({'ok':True, 'filename':fname}); return

        m = re.match(r'^/api/guards/([^/]+)/leave$', path)
        if m:
            s2 = self.require_admin('manager')
            if not s2: return
            for f in ('start_date','end_date'):
                if not data.get(f): self.err(f'{f} required'); return
            lid = str(uuid.uuid4()); db = get_db()
            avail_from = data.get('available_from','')
            db.execute('''INSERT INTO guard_leave (id,guard_id,leave_type,start_date,end_date,notes,available_from)
                          VALUES (?,?,?,?,?,?,?)''',
                       (lid, m.group(1), data.get('leave_type','Fixed Leave'),
                        data['start_date'], data['end_date'], data.get('notes',''), avail_from))
            audit(db, s, 'LEAVE_CREATE',
                  f'{m.group(1)} {data["start_date"]}~{data["end_date"]}'
                  + (f' (free from {avail_from})' if avail_from else '')); db.commit()
            row = R(db.execute('SELECT * FROM guard_leave WHERE id=?', (lid,)).fetchone())
            db.close(); self.send_json(row, 201); return

        if path == '/api/sites':
            s2 = self.require_admin('manager')
            if not s2: return
            if not data.get('name') or not data.get('client_name'):
                self.err('name and client_name required'); return
            sid = str(uuid.uuid4()); db = get_db()
            db.execute('''INSERT INTO sites
                (id,name,client_name,address,default_rate,contact_name,contact_phone,lat,lng,geofence_radius)
                VALUES (?,?,?,?,?,?,?,?,?,?)''',
                       (sid,data['name'],data['client_name'],data.get('address',''),
                        float(data.get('default_rate',0)),
                        data.get('contact_name',''),data.get('contact_phone',''),
                        data.get('lat'), data.get('lng'), int(data.get('geofence_radius') or 200)))
            db.commit()
            audit(db, s, 'SITE_CREATE', data['name']); db.commit()
            site = R(db.execute('SELECT * FROM sites WHERE id=?',(sid,)).fetchone()); db.close()
            self.send_json(site, 201); return

        if path == '/api/checkpoints':
            s2 = self.require_admin('manager')
            if not s2: return
            for f in ['site_id','name','lat','lng']:
                if data.get(f) is None: self.err(f'{f} required'); return
            cid = str(uuid.uuid4()); db = get_db()
            db.execute('''INSERT INTO checkpoints (id,site_id,name,lat,lng,radius_m,sort_order)
                          VALUES (?,?,?,?,?,?,?)''',
                       (cid, data['site_id'], data['name'], float(data['lat']), float(data['lng']),
                        int(data.get('radius_m') or 40), int(data.get('sort_order') or 0)))
            audit(db, s, 'CHECKPOINT_CREATE', f"{data['name']} (site={data['site_id']})"); db.commit()
            cp = R(db.execute('SELECT * FROM checkpoints WHERE id=?',(cid,)).fetchone()); db.close()
            self.send_json(cp, 201); return

        if path == '/api/shifts':
            s2 = self.require_admin('manager')
            if not s2: return
            # end_time is optional: an empty end_time means "Required" (open-ended,
            # e.g. hospitality security running until the venue closes) — the
            # guard's actual worked hours still come from their real clock-out,
            # not this scheduled field.
            for f in ['guard_id','site_id','shift_date','start_time']:
                if not data.get(f): self.err(f'{f} required'); return
            shid = str(uuid.uuid4()); db = get_db()
            db.execute('''INSERT INTO shifts (id,guard_id,site_id,shift_date,start_time,end_time,
                          position,notes,created_by,published) VALUES (?,?,?,?,?,?,?,?,?,0)''',
                       (shid, data['guard_id'], data['site_id'], data['shift_date'],
                        data['start_time'], data.get('end_time',''), data.get('position',''),
                        data.get('notes',''), s['name']))
            # Independent of whatever the modal warned about — a direct API
            # call wouldn't have shown that dialog at all — so this is the
            # authoritative record of the conflict, not just an echo of the UI.
            conflict = availability_conflict(db, data['guard_id'], data['shift_date'], data['start_time'])
            audit(db, s, 'SHIFT_CREATE',
                  f"{data['shift_date']} {data['start_time']}-{data.get('end_time') or 'Required'}"
                  + (f" (⚠ {conflict})" if conflict else '')); db.commit()
            row = with_shift_status([R(db.execute('''SELECT sh.*, g.name as guard_name, s.name as site_name
                FROM shifts sh JOIN guards g ON g.id=sh.guard_id JOIN sites s ON s.id=sh.site_id
                WHERE sh.id=?''', (shid,)).fetchone())])[0]
            db.close(); self.send_json(row, 201); return

        if path == '/api/shifts/publish':
            s2 = self.require_admin('manager')
            if not s2: return
            date_from = data.get('date_from'); date_to = data.get('date_to')
            if not date_from or not date_to: self.err('date_from and date_to required'); return
            db = get_db()
            rows = RL(db.execute('''
                SELECT sh.*, g.name as guard_name, g.email as guard_email, s.name as site_name
                FROM shifts sh JOIN guards g ON g.id=sh.guard_id JOIN sites s ON s.id=sh.site_id
                WHERE sh.shift_date>=? AND sh.shift_date<=? AND sh.published=0 AND sh.cancelled=0
            ''', (date_from, date_to)).fetchall())
            if rows:
                db.execute('''UPDATE shifts SET published=1
                              WHERE shift_date>=? AND shift_date<=? AND published=0 AND cancelled=0''',
                           (date_from, date_to))
                audit(db, s, 'SHIFT_PUBLISH', f'{date_from}~{date_to} ({len(rows)} shifts)')
                db.commit()
            db.close()
            by_guard = {}
            for r in rows:
                by_guard.setdefault(r['guard_id'], {'name':r['guard_name'],'email':r['guard_email'],'shifts':[]})
                by_guard[r['guard_id']]['shifts'].append(r)
            notified = 0
            for g in by_guard.values():
                if not g['email']: continue
                lines = [f"- {sh['shift_date']} {sh['start_time']}-{sh['end_time'] or 'Required'} at {sh['site_name']}"
                         for sh in g['shifts']]
                body = (f"Hi {g['name']},\n\nYour shifts for {date_from} to {date_to} have been "
                        f"published:\n\n" + "\n".join(lines) +
                        f"\n\nView your roster in the Guard Portal: {APP_URL}\n\n— {COMPANY_NAME}")
                if send_email(g['email'], f'{COMPANY_NAME}: New shifts published', body):
                    notified += 1
            self.send_json({'published': len(rows), 'notified': notified}); return

        if path == '/api/rates':
            s2 = self.require_admin('manager')
            if not s2: return
            if not data.get('guard_id') or not data.get('site_id'):
                self.err('guard_id and site_id required'); return
            db = get_db()
            db.execute('INSERT OR REPLACE INTO rates (guard_id,site_id,rate) VALUES (?,?,?)',
                       (data['guard_id'],data['site_id'],float(data.get('rate',0))))
            db.commit(); db.close(); self.send_json({'ok':True}); return

        if path == '/api/reminders':
            s2 = self.require_admin('manager')
            if not s2: return
            guard_ids = data.get('guard_ids',[])
            msg = data.get('message','Please submit your pending shift(s).')
            if not guard_ids: self.err('guard_ids required'); return
            db = get_db()
            for gid in guard_ids:
                db.execute('INSERT INTO reminders (id,guard_id,message) VALUES (?,?,?)',
                           (str(uuid.uuid4()),gid,msg))
            audit(db, s, 'REMINDERS_SENT', f'{len(guard_ids)} guards'); db.commit(); db.close()
            self.send_json({'ok':True,'sent':len(guard_ids)}); return

        if path == '/api/messages':
            s2 = self.require_admin('manager')
            if not s2: return
            if not data.get('guard_id') or not (data.get('body') or '').strip():
                self.err('guard_id and body required'); return
            db = get_db()
            guard = R(db.execute('SELECT id,name FROM guards WHERE id=?', (data['guard_id'],)).fetchone())
            if not guard: db.close(); self.err('Guard not found', 404); return
            mid = str(uuid.uuid4())
            db.execute('INSERT INTO messages (id,guard_id,sender,sender_name,body) VALUES (?,?,?,?,?)',
                       (mid, guard['id'], 'admin', s2['name'], data['body'].strip()))
            audit(db, s2, 'MESSAGE_SEND', f"to {guard['name']}"); db.commit(); db.close()
            self.send_json({'id':mid,'ok':True}, 201); return

        if path == '/api/messages/read':
            if not data.get('guard_id'): self.err('guard_id required'); return
            db = get_db()
            db.execute("UPDATE messages SET read_at=? WHERE guard_id=? AND sender='guard' AND read_at IS NULL",
                       (datetime.now().isoformat(), data['guard_id']))
            db.commit(); db.close()
            self.send_json({'ok':True}); return

        if path == '/api/faqs':
            s2 = self.require_admin('manager')
            if not s2: return
            for f in ['question','keywords','answer']:
                if not (data.get(f) or '').strip(): self.err(f'{f} required'); return
            db = get_db()
            fid = str(uuid.uuid4())
            n = db.execute('SELECT COUNT(*) FROM faqs').fetchone()[0]
            db.execute('INSERT INTO faqs (id,question,keywords,answer,sort_order) VALUES (?,?,?,?,?)',
                       (fid, data['question'].strip(), data['keywords'].strip(), data['answer'].strip(), n))
            audit(db, s2, 'FAQ_CREATE', data['question'].strip()); db.commit(); db.close()
            self.send_json({'id':fid,'ok':True}, 201); return

        if path == '/api/submissions/bulk':
            s2 = self.require_admin('manager')
            if not s2: return
            ids = data.get('ids',[]); action = data.get('action'); note = data.get('note','')
            if not ids or action not in ('approve','reject','delete'):
                self.err('ids and action required'); return
            if action == 'delete':
                # Same bar as the single delete — administrator and up.
                s3 = self.require_admin('administrator')
                if not s3: return
                db = get_db()
                n = delete_submissions(db, ids, s)
                db.commit(); db.close()
                self.send_json({'ok':True,'deleted':n}); return
            status = 'approved' if action == 'approve' else 'rejected'
            db = get_db()
            for sid in ids:
                db.execute('''UPDATE submissions SET status=?,admin_note=?,reviewed_by=?,reviewed_at=?
                              WHERE id=?''',
                           (status, note, s['name'], datetime.now().isoformat(), sid))
            audit(db, s, f'BULK_{action.upper()}', f'{len(ids)} submissions'); db.commit(); db.close()
            self.send_json({'ok':True,'updated':len(ids)}); return

        if path == '/api/logo':
            s2 = self.require_admin('manager')
            if not s2: return
            if not data.get('logo_b64') or not data.get('logo_ext'):
                self.err('logo_b64 and logo_ext required'); return
            ext = data['logo_ext'].lower().lstrip('.')
            if ext not in ['png','jpg','jpeg','gif','webp','svg']:
                self.err('Unsupported type'); return
            for oe in ['png','jpg','jpeg','gif','webp','svg']:
                op = os.path.join(UPLOADS_PATH,f'company_logo.{oe}')
                if os.path.exists(op): os.remove(op)
            with open(os.path.join(UPLOADS_PATH,f'company_logo.{ext}'),'wb') as f:
                f.write(base64.b64decode(data['logo_b64']))
            self.send_json({'ok':True}); return

        if path == '/api/admins':
            s2 = self.require_admin('administrator')
            if not s2: return
            requested_role = data.get('role','manager')
            # Administrator cannot create superadmin accounts
            if s2.get('role') == 'administrator' and requested_role == 'superadmin':
                self.err('Insufficient permissions to create superadmin', 403); return
            if not data.get('name') or not data.get('email') or not data.get('password'):
                self.err('name, email and password required'); return
            h, salt = hash_password(data['password'])
            aid = str(uuid.uuid4()); db = get_db()
            try:
                db.execute('''INSERT INTO admins (id,name,email,password_hash,salt,role,must_change_password)
                              VALUES (?,?,?,?,?,?,1)''',
                           (aid,data['name'],data['email'].lower(),h,salt,requested_role))
                if requested_role == 'client':
                    for site_id in (data.get('site_ids') or []):
                        db.execute('INSERT OR IGNORE INTO client_sites (admin_id,site_id) VALUES (?,?)',
                                   (aid, site_id))
                audit(db, s, 'ADMIN_CREATE', data['email']); db.commit()
                admin = R(db.execute(
                    'SELECT id,name,email,role,active,created_at FROM admins WHERE id=?',(aid,)).fetchone())
                db.close()
                # Send welcome email with temp credentials (optional — requires SMTP env vars)
                role_label = {'superadmin':'Super Admin','administrator':'Administrator',
                              'manager':'Manager','viewer':'Viewer'}.get(requested_role, requested_role)
                send_email(
                    data['email'].lower(),
                    f"Your {COMPANY_NAME} Account — Action Required",
                    f"Hi {data['name']},\n\n"
                    f"An admin account ({role_label}) has been created for you on the "
                    f"{COMPANY_NAME} Guard Management System.\n\n"
                    f"Login URL:          {APP_URL}\n"
                    f"Email:              {data['email']}\n"
                    f"Temporary password: {data['password']}\n\n"
                    f"IMPORTANT: You will be prompted to set your own password the first time you log in. "
                    f"Your temporary password will no longer work after that.\n\n"
                    f"— {COMPANY_NAME}"
                )
                self.send_json(admin, 201)
            except Exception as e:
                db.close(); self.err('Email already exists', 409)
            return

        if path == '/api/change-password':
            current = data.get('current_password','')
            new_pw  = data.get('new_password','')
            if not current or not new_pw: self.err('current and new password required'); return
            db = get_db()
            row = R(db.execute('SELECT * FROM admins WHERE id=?',(s['admin_id'],)).fetchone())
            if not verify_password(current, row['password_hash'], row['salt']):
                db.close(); self.err('Current password incorrect', 401); return
            h, salt = hash_password(new_pw)
            db.execute('UPDATE admins SET password_hash=?,salt=? WHERE id=?',
                       (h, salt, s['admin_id']))
            audit(db, s, 'PASSWORD_CHANGED'); db.commit(); db.close()
            self.send_json({'ok':True}); return

        self.send_response(404); self.end_headers()

    # ── PUT ────────────────────────────────────────────────────────────────────
    def do_PUT(self):
        path = urlparse(self.path).path
        data = self.read_json()

        s = self.require_admin()
        if s is None: return
        if s.get('must_change_password'):
            self.err('Please set your password before continuing', 403); return

        m = re.match(r'^/api/faqs/([^/]+)$', path)
        if m:
            s2 = self.require_admin('manager')
            if not s2: return
            db = get_db()
            existing = R(db.execute('SELECT id FROM faqs WHERE id=?', (m.group(1),)).fetchone())
            if not existing: db.close(); self.err('FAQ not found', 404); return
            updates = []; params = []
            for field in ('question','keywords','answer'):
                if field in data:
                    updates.append(f'{field}=?'); params.append(data[field])
            if 'active' in data:
                updates.append('active=?'); params.append(1 if data['active'] else 0)
            if updates:
                params.append(m.group(1))
                db.execute(f"UPDATE faqs SET {','.join(updates)} WHERE id=?", params)
                audit(db, s2, 'FAQ_UPDATE', m.group(1)); db.commit()
            db.close(); self.send_json({'ok':True}); return

        m = re.match(r'^/api/submissions/([^/]+)$', path)
        if m:
            s2 = self.require_admin('manager')
            if not s2: return
            db = get_db()
            updates = []; params = []
            # Status change fields
            if 'status' in data:
                updates += ['status=?','reviewed_by=?','reviewed_at=?']
                params  += [data['status'], s['name'], datetime.now().isoformat()]
            if 'admin_note' in data:
                updates.append('admin_note=?'); params.append(data['admin_note'])
            # Editable hours fields (manager only)
            if 'total_hours' in data:
                updates.append('total_hours=?'); params.append(float(data['total_hours']))
            if 'start_time' in data:
                updates.append('start_time=?'); params.append(data['start_time'])
            if 'end_time' in data:
                updates.append('end_time=?'); params.append(data['end_time'])
            if 'shift_date' in data:
                updates.append('shift_date=?'); params.append(data['shift_date'])
            if 'notes' in data:
                updates.append('notes=?'); params.append(data['notes'])
            if updates:
                params.append(m.group(1))
                db.execute(f"UPDATE submissions SET {','.join(updates)} WHERE id=?", params)
                is_edit = any(k in data for k in ('total_hours','start_time','end_time','shift_date'))
                audit(db, s, 'SUB_EDITED' if is_edit else f"SUB_{data.get('status','UPDATED').upper()}", m.group(1))
                db.commit()
            row = R(db.execute('''SELECT sub.*,g.name as guard_name,s.name as site_name,s.client_name
                FROM submissions sub JOIN guards g ON g.id=sub.guard_id
                JOIN sites s ON s.id=sub.site_id WHERE sub.id=?''', (m.group(1),)).fetchone())
            db.close(); self.send_json(row); return

        m = re.match(r'^/api/guards/([^/]+)$', path)
        if m:
            s2 = self.require_admin('manager')
            if not s2: return
            db = get_db()
            updates=[]; params=[]
            for f in ('name','license_number','phone','email','notes','license_state',
                      'license_expiry'):
                if f in data: updates.append(f'{f}=?'); params.append(data[f])
            if 'base_rate' in data:
                updates.append('base_rate=?'); params.append(float(data['base_rate']))
            for f in ('active','hide_on_schedule','no_license_required',
                      'license_reminder_days','license_critical'):
                if f in data: updates.append(f'{f}=?'); params.append(int(data[f]))
            if updates:
                params.append(m.group(1))
                try:
                    db.execute(f"UPDATE guards SET {','.join(updates)} WHERE id=?", params)
                except sqlite3.IntegrityError:
                    db.close(); self.err('Another guard already has that email address', 409); return
                audit(db, s, 'GUARD_UPDATE', m.group(1)); db.commit()
            g = no_secrets(R(db.execute('SELECT * FROM guards WHERE id=?',(m.group(1),)).fetchone()))
            db.close(); self.send_json(g); return

        m = re.match(r'^/api/guards/([^/]+)/compliance/([^/]+)$', path)
        if m:
            s2 = self.require_admin('manager')
            if not s2: return
            guard_id, item_id = m.group(1), m.group(2)
            fields = {}
            if 'checked' in data:          fields['checked']          = int(data['checked'])
            if 'reference_no' in data:     fields['reference_no']     = data['reference_no']
            if 'expiry_date' in data:      fields['expiry_date']      = data['expiry_date'] or None
            if 'reminder_days' in data:    fields['reminder_days']    = int(data['reminder_days'] or 60)
            if 'critical' in data:         fields['critical']         = int(data['critical'])
            if 'show_to_customer' in data: fields['show_to_customer'] = int(data['show_to_customer'])
            if not fields: self.err('No fields to update'); return
            db = get_db()
            cols = ','.join(fields.keys()); qs_ = ','.join('?'*len(fields))
            upd  = ','.join(f'{k}=excluded.{k}' for k in fields)
            db.execute(f'''INSERT INTO guard_compliance (id,guard_id,item_id,{cols})
                          VALUES (?,?,?,{qs_})
                          ON CONFLICT(guard_id,item_id) DO UPDATE SET {upd}''',
                       [str(uuid.uuid4()), guard_id, item_id] + list(fields.values()))
            db.commit(); db.close()
            self.send_json({'ok':True}); return

        m = re.match(r'^/api/sites/([^/]+)$', path)
        if m:
            s2 = self.require_admin('manager')
            if not s2: return
            db = get_db()
            db.execute('''UPDATE sites SET name=?,client_name=?,address=?,default_rate=?,
                          contact_name=?,contact_phone=?,active=?,lat=?,lng=?,geofence_radius=? WHERE id=?''',
                       (data.get('name'),data.get('client_name'),data.get('address'),
                        float(data.get('default_rate',0)),data.get('contact_name'),
                        data.get('contact_phone'),int(data.get('active',1)),
                        data.get('lat'), data.get('lng'), int(data.get('geofence_radius') or 200),
                        m.group(1)))
            db.commit()
            audit(db, s, 'SITE_UPDATE', data.get('name','')); db.commit()
            site = R(db.execute('SELECT * FROM sites WHERE id=?',(m.group(1),)).fetchone())
            db.close(); self.send_json(site); return

        m = re.match(r'^/api/checkpoints/([^/]+)$', path)
        if m:
            s2 = self.require_admin('manager')
            if not s2: return
            db = get_db()
            db.execute('''UPDATE checkpoints SET name=?,lat=?,lng=?,radius_m=?,sort_order=?,active=?
                          WHERE id=?''',
                       (data.get('name'), float(data.get('lat',0)), float(data.get('lng',0)),
                        int(data.get('radius_m') or 40), int(data.get('sort_order') or 0),
                        int(data.get('active',1)), m.group(1)))
            db.commit()
            cp = R(db.execute('SELECT * FROM checkpoints WHERE id=?',(m.group(1),)).fetchone())
            db.close(); self.send_json(cp); return

        m = re.match(r'^/api/shifts/([^/]+)$', path)
        if m:
            s2 = self.require_admin('manager')
            if not s2: return
            db = get_db()
            updates=[]; params=[]
            for f in ('guard_id','site_id','shift_date','start_time','end_time','position','notes'):
                if f in data: updates.append(f'{f}=?'); params.append(data[f])
            if 'cancelled' in data: updates.append('cancelled=?'); params.append(int(data['cancelled']))
            if updates:
                params.append(m.group(1))
                db.execute(f"UPDATE shifts SET {','.join(updates)} WHERE id=?", params)
                note = ''
                # Skip the check on a plain cancellation — "cancelled this
                # shift (⚠ conflict)" isn't meaningful once it's cancelled.
                if data.get('cancelled') != 1:
                    row0 = R(db.execute('SELECT guard_id,shift_date,start_time FROM shifts WHERE id=?',
                                        (m.group(1),)).fetchone())
                    if row0:
                        conflict = availability_conflict(db, row0['guard_id'], row0['shift_date'], row0['start_time'])
                        if conflict: note = f' (⚠ {conflict})'
                audit(db, s, 'SHIFT_UPDATE', m.group(1) + note); db.commit()
            row = R(db.execute('''SELECT sh.*, g.name as guard_name, s.name as site_name
                FROM shifts sh JOIN guards g ON g.id=sh.guard_id JOIN sites s ON s.id=sh.site_id
                WHERE sh.id=?''', (m.group(1),)).fetchone())
            db.close()
            self.send_json(with_shift_status([row])[0] if row else None); return

        m = re.match(r'^/api/incidents/([^/]+)$', path)
        if m:
            s2 = self.require_admin('manager')
            if not s2: return
            db = get_db()
            updates=[]; params=[]
            if 'status' in data:
                updates += ['status=?','reviewed_by=?','reviewed_at=?']
                params  += [data['status'], s['name'], datetime.now().isoformat()]
            if 'admin_note' in data:
                updates.append('admin_note=?'); params.append(data['admin_note'])
            if updates:
                params.append(m.group(1))
                db.execute(f"UPDATE incidents SET {','.join(updates)} WHERE id=?", params)
                audit(db, s, 'INCIDENT_UPDATE', m.group(1)); db.commit()
            row = R(db.execute('''SELECT i.*,g.name as guard_name,s.name as site_name
                FROM incidents i JOIN guards g ON g.id=i.guard_id
                JOIN sites s ON s.id=i.site_id WHERE i.id=?''', (m.group(1),)).fetchone())
            db.close(); self.send_json(row); return

        m = re.match(r'^/api/admins/([^/]+)$', path)
        if m:
            s2 = self.require_admin('administrator')
            if not s2: return
            db = get_db()
            target = R(db.execute('SELECT role FROM admins WHERE id=?',(m.group(1),)).fetchone())
            # Administrator cannot modify superadmin accounts or assign superadmin role
            if s2.get('role') == 'administrator':
                if target and target.get('role') == 'superadmin':
                    db.close(); self.err('Insufficient permissions to modify superadmin', 403); return
                if data.get('role') == 'superadmin':
                    db.close(); self.err('Insufficient permissions to assign superadmin role', 403); return
            updates = []
            params  = []
            if 'role'   in data: updates.append('role=?');   params.append(data['role'])
            if 'active' in data: updates.append('active=?'); params.append(int(data['active']))
            if 'name'   in data: updates.append('name=?');   params.append(data['name'])
            if data.get('new_password'):
                h, salt = hash_password(data['new_password'])
                updates += ['password_hash=?','salt=?']; params += [h, salt]
            if updates:
                params.append(m.group(1))
                db.execute(f"UPDATE admins SET {','.join(updates)} WHERE id=?", params)
                audit(db, s, 'ADMIN_UPDATE', m.group(1)); db.commit()
            if 'site_ids' in data:
                db.execute('DELETE FROM client_sites WHERE admin_id=?', (m.group(1),))
                for site_id in (data.get('site_ids') or []):
                    db.execute('INSERT OR IGNORE INTO client_sites (admin_id,site_id) VALUES (?,?)',
                               (m.group(1), site_id))
                db.commit()
            row = R(db.execute(
                'SELECT id,name,email,role,active,last_login,created_at FROM admins WHERE id=?',
                (m.group(1),)).fetchone())
            if row and row['role'] == 'client':
                row['site_ids'] = self.client_site_ids(m.group(1))
            db.close(); self.send_json(row); return

        self.send_response(404); self.end_headers()

    # ── DELETE ─────────────────────────────────────────────────────────────────
    def do_DELETE(self):
        path = urlparse(self.path).path

        # Admin delete — superadmin only, cannot delete self
        m = re.match(r'^/api/admins/([^/]+)$', path)
        if m:
            s = self.require_admin('superadmin')
            if s is None: return
            if m.group(1) == s.get('admin_id'):
                self.err('Cannot delete your own account', 400); return
            db = get_db()
            target = R(db.execute('SELECT id,name FROM admins WHERE id=?',(m.group(1),)).fetchone())
            if not target:
                db.close(); self.err('Admin not found', 404); return
            db.execute('DELETE FROM admins WHERE id=?', (m.group(1),))
            audit(db, s, 'ADMIN_DELETE', target.get('name',m.group(1))); db.commit(); db.close()
            self.send_json({'ok':True}); return

        s = self.require_admin('manager')
        if s is None: return
        m = re.match(r'^/api/rates/([^/]+)/([^/]+)$', path)
        if m:
            db = get_db()
            db.execute('DELETE FROM rates WHERE guard_id=? AND site_id=?', (m.group(1),m.group(2)))
            db.commit(); db.close(); self.send_json({'ok':True}); return
        m = re.match(r'^/api/checkpoints/([^/]+)$', path)
        if m:
            db = get_db()
            db.execute('DELETE FROM checkpoints WHERE id=?', (m.group(1),))
            audit(db, s, 'CHECKPOINT_DELETE', m.group(1)); db.commit(); db.close()
            self.send_json({'ok':True}); return
        m = re.match(r'^/api/leave/([^/]+)$', path)
        if m:
            db = get_db()
            db.execute('DELETE FROM guard_leave WHERE id=?', (m.group(1),))
            audit(db, s, 'LEAVE_DELETE', m.group(1)); db.commit(); db.close()
            self.send_json({'ok':True}); return
        m = re.match(r'^/api/faqs/([^/]+)$', path)
        if m:
            db = get_db()
            db.execute('DELETE FROM faqs WHERE id=?', (m.group(1),))
            audit(db, s, 'FAQ_DELETE', m.group(1)); db.commit(); db.close()
            self.send_json({'ok':True}); return
        m = re.match(r'^/api/shifts/([^/]+)$', path)
        if m:
            db = get_db()
            db.execute('DELETE FROM shifts WHERE id=?', (m.group(1),))
            audit(db, s, 'SHIFT_DELETE', m.group(1)); db.commit(); db.close()
            self.send_json({'ok':True}); return
        m = re.match(r'^/api/submissions/([^/]+)$', path)
        if m:
            # Administrator and up only: submissions are the pay/invoice record,
            # so removing one is a heavier action than the approve/reject a
            # manager can do.
            s2 = self.require_admin('administrator')
            if not s2: return
            db = get_db()
            deleted = delete_submissions(db, [m.group(1)], s)
            db.commit(); db.close()
            if not deleted: self.err('Submission not found', 404); return
            self.send_json({'ok':True,'deleted':deleted}); return
        self.send_response(404); self.end_headers()

    def do_OPTIONS(self):
        # The frontend is served from this same server, so it never makes a
        # cross-origin request and never needs CORS headers — the previous
        # wildcard Access-Control-Allow-Origin was broader than the app
        # actually requires. (In practice it was already inert: none of the
        # real GET/POST/PUT/DELETE responses ever sent that header either, so
        # a genuinely cross-origin caller could pass preflight but still never
        # read a response — this just removes the unnecessary opening cleanly.)
        self.send_response(204)
        self.security_headers()
        self.end_headers()

# ─── Main ─────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    init_db()
    try:
        run_db_backup()
    except Exception as e:
        print(f'  BACKUP: startup snapshot failed: {e}')
    threading.Thread(target=_backup_loop, daemon=True).start()
    print(f"\n{'='*55}")
    print(f"  {COMPANY_NAME}")
    print(f"  Guard Management System  v3.0")
    print(f"{'='*55}")
    print(f"\n  http://localhost:{PORT}")
    print(f"  Admin email:    {DEFAULT_ADMIN_EMAIL}")
    print(f"  Admin password: {DEFAULT_ADMIN_PASSWORD}")
    print(f"\n  Press Ctrl+C to stop\n{'='*55}\n")
    with http.server.HTTPServer(('', PORT), Handler) as srv:
        try: srv.serve_forever()
        except KeyboardInterrupt: print('\nStopped.')
